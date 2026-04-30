from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
import re
from typing import Any

from openpyxl import load_workbook


@dataclass
class AforParsedRow:
    row_index: int
    status: str
    errors: list[str]
    data: dict[str, Any]


ALARM_LEVEL_MAP = {
    "1ST": "First Alarm",
    "1ST ALARM": "First Alarm",
    "FIRST": "First Alarm",
    "FIRST ALARM": "First Alarm",
    "2ND": "Second Alarm",
    "2ND ALARM": "Second Alarm",
    "SECOND": "Second Alarm",
    "SECOND ALARM": "Second Alarm",
    "3RD": "Third Alarm",
    "3RD ALARM": "Third Alarm",
    "THIRD": "Third Alarm",
    "THIRD ALARM": "Third Alarm",
    "4TH": "Fourth Alarm",
    "4TH ALARM": "Fourth Alarm",
    "FOURTH": "Fourth Alarm",
    "FOURTH ALARM": "Fourth Alarm",
    "5TH": "Fifth Alarm",
    "5TH ALARM": "Fifth Alarm",
    "FIFTH": "Fifth Alarm",
    "FIFTH ALARM": "Fifth Alarm",
    "TF ALPHA": "Task Force Alpha",
    "TASK FORCE ALPHA": "Task Force Alpha",
    "TF BRAVO": "Task Force Bravo",
    "TASK FORCE BRAVO": "Task Force Bravo",
    "TF CHARLIE": "Task Force Charlie",
    "TASK FORCE CHARLIE": "Task Force Charlie",
    "TF DELTA": "Task Force Delta",
    "TASK FORCE DELTA": "Task Force Delta",
    "GENERAL": "General Alarm",
    "GENERAL ALARM": "General Alarm",
}


def _safe_int(val: Any, default: int = 0) -> int:
    if val is None or val == "" or val == "N/A":
        return default
    try:
        if isinstance(val, (int, float)):
            return int(val)
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return default


def _safe_float(val: Any, default: float = 0.0) -> float:
    if val is None or val == "" or val == "N/A":
        return default
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return default


def _safe_dt(val: Any) -> str | None:
    if isinstance(val, datetime):
        return val.isoformat()
    if not val:
        return None

    if isinstance(val, (int, float)):
        try:
            serial = float(val)
            base = datetime(1899, 12, 30)
            dt = base + timedelta(days=serial)
            if serial < 1:
                return dt.strftime("%H:%M:%S")
            return dt.isoformat()
        except Exception:
            return None

    raw_numeric = str(val).strip()
    try:
        serial = float(raw_numeric)
        base = datetime(1899, 12, 30)
        dt = base + timedelta(days=serial)
        if serial < 1:
            return dt.strftime("%H:%M:%S")
        return dt.isoformat()
    except (ValueError, TypeError):
        pass

    dt_str = str(val).strip()
    for fmt in [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%m-%d-%Y %H:%M:%S",
        "%m-%d-%Y %H:%M",
        "%H:%M",
        "%H:%M:%S",
        "%Y-%m-%d",
        "%m-%d-%Y",
        "%m/%d/%Y",
    ]:
        try:
            return datetime.strptime(dt_str, fmt).isoformat()
        except ValueError:
            continue
    return None


def _parse_area_pair(val: Any) -> tuple[float, float]:
    if val is None:
        return 0.0, 0.0
    s = str(val).strip()
    if "/" in s:
        parts = s.split("/", 1)
        return _safe_float(parts[0].strip()), _safe_float(parts[1].strip())
    return _safe_float(s), 0.0


_COORD_RE = re.compile(r"^([A-Z]+)(\d+)$")


def _cell_str(ws: Any, coord: str) -> str:
    try:
        v = ws[coord].value
    except Exception:
        return ""
    if v is None:
        return ""
    return str(v).strip()


def _find_structural_marker_rows(ws: Any) -> tuple[int | None, int | None]:
    title_row: int | None = None
    section_row: int | None = None

    for row in range(1, 161):
        row_values = [_cell_str(ws, f"{col}{row}").upper() for col in ("A", "B", "C", "D", "E", "F")]
        combined = " ".join(v for v in row_values if v).strip()

        if title_row is None and "AFTER FIRE OPERATIONS REPORT" in combined:
            title_row = row
        if section_row is None and "A. RESPONSE DETAILS" in combined:
            section_row = row

        if title_row is not None and section_row is not None:
            break

    return title_row, section_row


def _combine_date_and_time(notification_dt: str | None, time_value: Any) -> str | None:
    if not notification_dt or not time_value:
        return None

    date_part = str(notification_dt).split("T", 1)[0]
    return _safe_dt(f"{date_part} {str(time_value).strip()}")


class BfpXlsxParser:
    def __init__(self, ws):
        self.ws = ws
        self._row_offset = self._infer_row_offset()

    def _infer_row_offset(self) -> int:
        title_row, section_row = _find_structural_marker_rows(self.ws)
        if title_row is None:
            return 0

        offset = title_row - 14
        if section_row is not None and (section_row - 18) != offset:
            return 0
        return offset

    def _coord_with_offset(self, coord: str) -> str:
        match = _COORD_RE.match(coord.upper())
        if not match or self._row_offset == 0:
            return coord

        col, row_str = match.groups()
        shifted_row = max(1, int(row_str) + self._row_offset)
        return f"{col}{shifted_row}"

    def get(self, coord: str) -> Any:
        shifted_coord = self._coord_with_offset(coord)
        val = self.ws[shifted_coord].value
        if val is None and shifted_coord != coord:
            val = self.ws[coord].value
        if val is None:
            return None
        if isinstance(val, str):
            return val.strip()
        return val

    def _is_marked(self, coord: str) -> bool:
        raw = self.get(coord)
        if raw is None:
            return False

        if isinstance(raw, bool):
            return raw

        if isinstance(raw, (int, float)):
            return raw != 0

        val = str(raw).strip().lower()
        if not val:
            return False

        if val.startswith("="):
            expr = val.lstrip("=").strip().lower()
            if expr in {"true", "1"}:
                return True

        return val in {
            "x",
            "1",
            "true",
            "v",
            "/",
            "yes",
            "checked",
            "☑",
            "☒",
            "✓",
            "✔",
            "✅",
        }

    def _is_marked_on_row(self, row: int, cols: tuple[str, ...] = ("B", "C", "D")) -> bool:
        return any(self._is_marked(f"{col}{row}") for col in cols)

    def _male_female_pair(self, row: int) -> tuple[Any, Any]:
        candidate_pairs = [("D", "E"), ("C", "D"), ("E", "F"), ("F", "G")]
        fallback_pair = (None, None)
        for male_col, female_col in candidate_pairs:
            male_val = self.get(f"{male_col}{row}")
            female_val = self.get(f"{female_col}{row}")
            if fallback_pair == (None, None):
                fallback_pair = (male_val, female_val)

            has_male = male_val not in (None, "")
            has_female = female_val not in (None, "")
            if has_male or has_female:
                return male_val, female_val

        return fallback_pair

    def parse(self) -> dict[str, Any]:
        responder_type = (
            "First Responder"
            if self._is_marked("B20")
            else ("Augmenting Team" if self._is_marked("B21") else "First Responder")
        )

        def _na(v: Any) -> bool:
            return v is None or str(v).strip().upper() in ("", "N/A")

        classification = "Structural"
        cat_val = None
        if self._is_marked_on_row(48):
            classification = "Structural"
            raw48 = self.get("D48")
            cat_val = None if _na(raw48) else raw48
        elif self._is_marked_on_row(49):
            classification = "Non-Structural"
            raw49 = self.get("D49")
            cat_val = None if _na(raw49) else raw49
        elif self._is_marked_on_row(50):
            classification = "Transportation"
            raw50 = self.get("D50")
            cat_val = None if _na(raw50) else raw50
        else:
            d48, d49, d50 = self.get("D48"), self.get("D49"), self.get("D50")
            if not _na(d48):
                classification = "Structural"
                cat_val = d48
            elif not _na(d49):
                classification = "Non-Structural"
                cat_val = d49
            elif not _na(d50):
                classification = "Transportation"
                cat_val = d50

        stage = self.get("D54") or self.get("B54")
        if stage and "pick from dropdown" in str(stage).lower():
            stage = None

        extent = "None/Minor Damage"
        extent_dval: Any = None
        extent_map = {
            56: "None/Minor Damage",
            57: "Confined to Object/Vehicle",
            58: "Confined to Room",
            59: "Confined to Structure or Property",
            60: "Total Loss",
            61: "Extended Beyond Structure or Property",
        }
        for ext_row, ext_label in extent_map.items():
            if self._is_marked_on_row(ext_row):
                extent = ext_label
                raw_d = self.get(f"D{ext_row}")
                extent_dval = None if (raw_d is None or str(raw_d).strip().upper() == "N/A") else raw_d
                break
        else:
            for ext_row in extent_map:
                raw_d = self.get(f"D{ext_row}")
                if raw_d is not None and str(raw_d).strip().upper() not in ("", "N/A"):
                    extent_dval = raw_d
                    break

        problems = []
        prob_map = {
            "B195": "Inaccurate address",
            "B196": "Geographically challenged",
            "B197": "Road conditions",
            "B198": "Road under construction",
            "B199": "Traffic congestion",
            "B200": "Road accidents",
            "B201": "Vehicles failure to yield",
            "B202": "Natural Disasters",
            "B203": "Civil Disturbance",
            "B204": "Uncooperative or panicked residents",
            "B205": "Safety and security threats",
            "B206": "Property security or owner delays",
            "B207": "Engine failure",
            "B208": "Uncooperative fire auxiliary",
            "B209": "Poor water supply access",
            "B210": "Intense heat and smoke",
            "B211": "Structural hazards",
            "B212": "Equipment malfunction",
            "B213": "Poor inter-agency coordination",
            "B214": "Radio communication breakdown",
            "B215": "HazMat risks",
            "B216": "Physical exhaustion and injuries",
            "B217": "Emotional and psychological effects",
            "B218": "Community complaints",
            "B219": "Others",
        }
        for c, flavor in prob_map.items():
            row_num = int(c[1:])
            if self._is_marked_on_row(row_num):
                problems.append(flavor)

        icp_present = self._is_marked_on_row(102)
        icp_location = self.get("D102") if icp_present else None

        narrative_lines = []
        for r in range(160, 191):
            line = self.get(f"B{r}")
            if line:
                narrative_lines.append(str(line))

        others = []
        for r in range(124, 133):
            name = self.get(f"B{r}")
            rem = self.get(f"E{r}")
            if name and "N/A" not in str(name).upper():
                others.append({"name": name, "designation": rem or ""})

        inj_civ_m, inj_civ_f = self._male_female_pair(106)
        inj_bfp_m, inj_bfp_f = self._male_female_pair(107)
        inj_aux_m, inj_aux_f = self._male_female_pair(108)
        fat_civ_m, fat_civ_f = self._male_female_pair(109)
        fat_bfp_m, fat_bfp_f = self._male_female_pair(110)
        fat_aux_m, fat_aux_f = self._male_female_pair(111)

        engines = [self.get("D31"), self.get("D32"), self.get("D33")]
        engine_list = [
            str(e).strip()
            for e in engines
            if e is not None and str(e).strip() and str(e).strip().upper() != "N/A"
        ]

        return {
            "responder_type": responder_type,
            "fire_station_name": self.get("D20") if responder_type == "First Responder" else self.get("D21"),
            "notification_date": self.get("D22"),
            "notification_time": self.get("D23"),
            "region": self.get("D24"),
            "province": self.get("D25"),
            "city": self.get("D26"),
            "address": self.get("D27"),
            "landmark": self.get("D28"),
            "caller_info": self.get("D29"),
            "receiver": self.get("D30"),
            "engine": ", ".join(engine_list),
            "time_dispatched": self.get("D34"),
            "time_arrived": self.get("D37"),
            "response_time": self.get("D40"),
            "distance_km": self.get("D41"),
            "alarm_level": self.get("D42"),
            "time_returned": self.get("D43"),
            "gas_liters": self.get("D44"),
            "classification": classification,
            "category": cat_val,
            "owner": self.get("D51"),
            "description": self.get("D52"),
            "origin": self.get("D53"),
            "stage": stage,
            "extent": extent,
            "extent_total_floor_area_sqm": extent_dval,
            "extent_total_land_area_hectares": extent_dval,
            "struct_aff": self.get("D62"),
            "house_aff": self.get("D63"),
            "fam_aff": self.get("D64"),
            "indiv_aff": self.get("D65"),
            "vehic_aff": self.get("D66"),
            "res_bfp_truck": self.get("D70"),
            "res_lgu_truck": self.get("D71"),
            "res_vol_truck": self.get("D72"),
            "res_bfp_amb": self.get("D73"),
            "res_non_amb": self.get("D74"),
            "res_bfp_resc": self.get("D75"),
            "res_non_resc": self.get("D76"),
            "res_others": self.get("D77"),
            "tool_scba": self.get("D79"),
            "tool_rope": self.get("D80"),
            "tool_ladder": self.get("D81"),
            "tool_hose": self.get("D82"),
            "tool_hydra": self.get("D83"),
            "tool_others": self.get("D84"),
            "hydrant_dist": self.get("D85"),
            "timeline": {
                "foua": {
                    "time": self.get("D88"),
                    "date": self.get("E88"),
                    "commander": self.get("F88"),
                },
                "alarm_1st": {"time": self.get("D89"), "date": self.get("E89"), "commander": self.get("F89")},
                "alarm_2nd": {"time": self.get("D90"), "date": self.get("E90"), "commander": self.get("F90")},
                "alarm_3rd": {"time": self.get("D91"), "date": self.get("E91"), "commander": self.get("F91")},
                "alarm_4th": {"time": self.get("D92"), "date": self.get("E92"), "commander": self.get("F92")},
                "alarm_5th": {"time": self.get("D93"), "date": self.get("E93"), "commander": self.get("F93")},
                "tf_alpha": {"time": self.get("D94"), "date": self.get("E94"), "commander": self.get("F94")},
                "tf_bravo": {"time": self.get("D95"), "date": self.get("E95"), "commander": self.get("F95")},
                "tf_charlie": {"time": self.get("D96"), "date": self.get("E96"), "commander": self.get("F96")},
                "tf_delta": {"time": self.get("D97"), "date": self.get("E97"), "commander": self.get("F97")},
                "general": {"time": self.get("D98"), "date": self.get("E98"), "commander": self.get("F98")},
                "fuc": {"time": self.get("D99"), "date": self.get("E99"), "commander": self.get("F99")},
                "fo": {"time": self.get("D100"), "date": self.get("E100"), "commander": self.get("F100")},
            },
            "icp_present": icp_present,
            "icp_location": icp_location,
            "inj_civ_m": inj_civ_m,
            "inj_civ_f": inj_civ_f,
            "inj_bfp_m": inj_bfp_m,
            "inj_bfp_f": inj_bfp_f,
            "inj_aux_m": inj_aux_m,
            "inj_aux_f": inj_aux_f,
            "fat_civ_m": fat_civ_m,
            "fat_civ_f": fat_civ_f,
            "fat_bfp_m": fat_bfp_m,
            "fat_bfp_f": fat_bfp_f,
            "fat_aux_m": fat_aux_m,
            "fat_aux_f": fat_aux_f,
            "pod_commander": self.get("D114"),
            "pod_shift": self.get("D115"),
            "pod_nozzleman": self.get("D116"),
            "pod_lineman": self.get("D117"),
            "pod_crew": self.get("D118"),
            "pod_dpo": self.get("D119"),
            "pod_safety": self.get("D120"),
            "pod_inv": self.get("D121"),
            "others_list": others,
            "narrative": "\n".join(narrative_lines),
            "problems": problems,
            "recommendations": self.get("B222"),
            "disposition": self.get("B229"),
            "prepared_by": self.get("C238"),
            "noted_by": self.get("F238"),
        }


def parse_afor_report_data(data: dict, region_id: int) -> AforParsedRow:
    errors: list[str] = []

    def _dt(d: Any, t: Any = None) -> str | None:
        if not d:
            return None

        if t:
            if isinstance(d, datetime) and hasattr(t, "hour") and hasattr(t, "minute"):
                try:
                    return datetime.combine(d.date(), t).isoformat()
                except Exception:
                    pass

            d_serial: float | None = None
            t_serial: float | None = None
            try:
                d_serial = float(d)
                t_serial = float(t)
            except (TypeError, ValueError):
                d_serial = None
                t_serial = None

            if d_serial is not None and t_serial is not None:
                try:
                    base = datetime(1899, 12, 30)
                    date_dt = base + timedelta(days=d_serial)
                    time_dt = base + timedelta(days=t_serial)
                    merged = datetime.combine(date_dt.date(), time_dt.time())
                    return merged.isoformat()
                except Exception:
                    pass

            date_part = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d).split(" ")[0]
            return _safe_dt(f"{date_part} {str(t).strip()}")

        return _safe_dt(d)

    notif_dt = _dt(data.get("notification_date"), data.get("notification_time"))

    ci = str(data.get("caller_info") or "")
    c_name = ci.split("/")[0].strip() if "/" in ci else ci
    c_num = ci.split("/")[1].strip() if "/" in ci else ""

    timeline = data.get("timeline") or {
        "foua": {"time": None, "date": data.get("notification_date"), "commander": None},
        "alarm_1st": {"time": data.get("alarm_1st"), "date": data.get("notification_date")},
        "alarm_2nd": {"time": None, "date": data.get("notification_date"), "commander": None},
        "alarm_3rd": {"time": None, "date": data.get("notification_date"), "commander": None},
        "alarm_4th": {"time": None, "date": data.get("notification_date"), "commander": None},
        "alarm_5th": {"time": None, "date": data.get("notification_date"), "commander": None},
        "tf_alpha": {"time": None, "date": data.get("notification_date"), "commander": None},
        "tf_bravo": {"time": None, "date": data.get("notification_date"), "commander": None},
        "tf_charlie": {"time": None, "date": data.get("notification_date"), "commander": None},
        "tf_delta": {"time": None, "date": data.get("notification_date"), "commander": None},
        "general": {"time": None, "date": data.get("notification_date"), "commander": None},
        "fuc": {"time": None, "date": data.get("notification_date"), "commander": None},
        "fo": {"time": None, "date": data.get("notification_date"), "commander": None},
    }

    pod_safety_raw = str(data.get("pod_safety") or "").strip()
    if "/" in pod_safety_raw:
        pod_safety_name, pod_safety_contact = [s.strip() for s in pod_safety_raw.split("/", 1)]
    else:
        pod_safety_name, pod_safety_contact = pod_safety_raw, ""

    pod_inv_raw = str(data.get("pod_inv") or "").strip()
    if "/" in pod_inv_raw:
        pod_inv_name, pod_inv_contact = [s.strip() for s in pod_inv_raw.split("/", 1)]
    else:
        pod_inv_name, pod_inv_contact = pod_inv_raw, ""

    engine_dispatched = data.get("engine") or data.get("engine_dispatched") or ""
    time_engine_dispatched = _safe_dt(data.get("time_dispatched") or data.get("time_engine_dispatched"))
    time_arrived_at_scene = _safe_dt(data.get("time_arrived") or data.get("time_arrived_at_scene"))
    time_returned_to_base = _safe_dt(data.get("time_returned") or data.get("time_returned_to_base"))
    receiver_name_val = data.get("receiver") or data.get("receiver_name") or ""

    classification_of_involved = data.get("classification") or data.get("classification_of_involved") or ""
    type_of_involved_general_category = data.get("category") or data.get("type_of_involved_general_category") or ""
    general_description_of_involved = data.get("description") or data.get("general_description_of_involved") or ""
    stage_of_fire_upon_arrival = data.get("stage") or data.get("stage_of_fire_upon_arrival") or ""

    raw_area = data.get("extent_total_floor_area_sqm")
    if raw_area is not None and "/" in str(raw_area):
        floor_sqm, land_ha = _parse_area_pair(raw_area)
    else:
        floor_sqm = _safe_float(raw_area)
        land_ha = _safe_float(data.get("extent_total_land_area_hectares"))

    ns = {
        "notification_dt": notif_dt,
        "responder_type": data.get("responder_type"),
        "fire_station_name": data.get("fire_station_name") or "",
        "alarm_level": ALARM_LEVEL_MAP.get(str(data.get("alarm_level") or "").strip().upper(), data.get("alarm_level")),
        "engine_dispatched": engine_dispatched,
        "time_engine_dispatched": time_engine_dispatched or "",
        "time_arrived_at_scene": time_arrived_at_scene or "",
        "time_returned_to_base": time_returned_to_base or "",
        "distance_to_fire_scene_km": _safe_float(data.get("distance_km") if data.get("distance_km") is not None else data.get("distance_to_fire_scene_km")),
        "receiver_name": receiver_name_val,
        "general_category": classification_of_involved,
        "sub_category": type_of_involved_general_category,
        "classification_of_involved": classification_of_involved,
        "type_of_involved_general_category": type_of_involved_general_category,
        "general_description_of_involved": general_description_of_involved,
        "fire_origin": data.get("origin") or data.get("area_of_origin"),
        "extent_of_damage": data.get("extent") or data.get("extent_of_damage"),
        "stage_of_fire": stage_of_fire_upon_arrival,
        "stage_of_fire_upon_arrival": stage_of_fire_upon_arrival,
        "structures_affected": _safe_int(data.get("struct_aff") if data.get("struct_aff") is not None else data.get("structures_affected")),
        "households_affected": _safe_int(data.get("house_aff")),
        "families_affected": _safe_int(data.get("fam_aff")),
        "individuals_affected": _safe_int(data.get("indiv_aff")),
        "vehicles_affected": _safe_int(data.get("vehic_aff")),
        "distance_from_station_km": _safe_float(data.get("distance_km") if data.get("distance_km") is not None else data.get("distance_from_station_km")),
        "total_response_time_minutes": _safe_int(data.get("response_time")),
        "total_gas_consumed_liters": _safe_float(data.get("gas_liters")),
        "extent_total_floor_area_sqm": floor_sqm,
        "extent_total_land_area_hectares": land_ha,
        "resources_deployed": {
            "bfp_fire_trucks": _safe_int(data.get("res_bfp_truck") if data.get("res_bfp_truck") is not None else data.get("res_bfp_trucks")),
            "lgu_owned_trucks": _safe_int(data.get("res_lgu_truck")),
            "non_bfp_trucks": _safe_int(data.get("res_vol_truck")),
            "bfp_ambulance": _safe_int(data.get("res_bfp_amb")),
            "non_bfp_ambulance": _safe_int(data.get("res_non_amb")),
            "bfp_rescue_trucks": _safe_int(data.get("res_bfp_resc")),
            "non_bfp_rescue_trucks": _safe_int(data.get("res_non_resc")),
            "other_vehicles": str(data.get("res_others") or ""),
            "scba": _safe_int(data.get("tool_scba")),
            "rope": _safe_int(data.get("tool_rope")),
            "ladder": _safe_int(data.get("tool_ladder")),
            "hoseline": _safe_int(data.get("tool_hose")),
            "hydrolic_tools": _safe_int(data.get("tool_hydra")),
            "other_tools": str(data.get("tool_others") or ""),
            "hydrant_distance": str(data.get("hydrant_dist") or ""),
        },
        "alarm_timeline": {
            "alarm_foua": _dt(timeline["foua"]["date"], timeline["foua"]["time"]),
            "alarm_foua_commander": timeline["foua"].get("commander"),
            "alarm_1st": _dt(timeline["alarm_1st"]["date"], timeline["alarm_1st"]["time"]),
            "alarm_1st_commander": timeline["alarm_1st"].get("commander"),
            "alarm_2nd": _dt(timeline["alarm_2nd"]["date"], timeline["alarm_2nd"]["time"]),
            "alarm_2nd_commander": timeline["alarm_2nd"].get("commander"),
            "alarm_3rd": _dt(timeline["alarm_3rd"]["date"], timeline["alarm_3rd"]["time"]),
            "alarm_3rd_commander": timeline["alarm_3rd"].get("commander"),
            "alarm_4th": _dt(timeline["alarm_4th"]["date"], timeline["alarm_4th"]["time"]),
            "alarm_4th_commander": timeline["alarm_4th"].get("commander"),
            "alarm_5th": _dt(timeline["alarm_5th"]["date"], timeline["alarm_5th"]["time"]),
            "alarm_5th_commander": timeline["alarm_5th"].get("commander"),
            "alarm_tf_alpha": _dt(timeline["tf_alpha"]["date"], timeline["tf_alpha"]["time"]),
            "alarm_tf_alpha_commander": timeline["tf_alpha"].get("commander"),
            "alarm_tf_bravo": _dt(timeline["tf_bravo"]["date"], timeline["tf_bravo"]["time"]),
            "alarm_tf_bravo_commander": timeline["tf_bravo"].get("commander"),
            "alarm_tf_charlie": _dt(timeline["tf_charlie"]["date"], timeline["tf_charlie"]["time"]),
            "alarm_tf_charlie_commander": timeline["tf_charlie"].get("commander"),
            "alarm_tf_delta": _dt(timeline["tf_delta"]["date"], timeline["tf_delta"]["time"]),
            "alarm_tf_delta_commander": timeline["tf_delta"].get("commander"),
            "alarm_general": _dt(timeline["general"]["date"], timeline["general"]["time"]),
            "alarm_general_commander": timeline["general"].get("commander"),
            "alarm_fuc": _dt(timeline["fuc"]["date"], timeline["fuc"]["time"]),
            "alarm_fuc_commander": timeline["fuc"].get("commander"),
            "alarm_fo": _dt(timeline["fo"]["date"], timeline["fo"]["time"]),
            "alarm_fo_commander": timeline["fo"].get("commander"),
        },
        "problems_encountered": data.get("problems", []),
        "recommendations": data.get("recommendations") or "",
        "region": data.get("region") or "",
        "province_district": data.get("province") or "",
        "city_municipality": data.get("city") or "",
        "incident_address": data.get("address") or "",
        "nearest_landmark": data.get("landmark") or "",
    }

    mapped = {
        "region_id": region_id,
        "caller_info": f"{c_name} / {c_num}" if c_name and c_num else (c_name or c_num),
        "incident_nonsensitive_details": ns,
        "incident_sensitive_details": {
            "caller_name": c_name,
            "caller_number": c_num,
            "receiver_name": receiver_name_val,
            "owner_name": data.get("owner") or "",
            "establishment_name": data.get("owner") or "",
            "street_address": data.get("address") or "",
            "landmark": data.get("landmark") or "",
            "personnel_on_duty": {
                "engine_commander": data.get("pod_commander") or "",
                "shift_in_charge": data.get("pod_shift") or "",
                "nozzleman": data.get("pod_nozzleman") or "",
                "lineman": data.get("pod_lineman") or "",
                "engine_crew": data.get("pod_crew") or "",
                "driver": data.get("pod_dpo") or "",
                "pump_operator": data.get("pod_dpo") or "",
                "safety_officer": {"name": pod_safety_name, "contact": pod_safety_contact},
                "fire_arson_investigator": {"name": pod_inv_name, "contact": pod_inv_contact},
            },
            "other_personnel": data.get("others_list", []),
            "narrative_report": data.get("narrative") or "",
            "is_icp_present": bool(data.get("icp_present")),
            "icp_location": data.get("icp_location") or "",
        },
        "_city_text": data.get("city") or "",
        "_province_text": data.get("province") or "",
        "_region_text": data.get("region") or "",
        "_form_kind": "STRUCTURAL_AFOR",
    }

    if not notif_dt:
        errors.append("Missing required fields: notification_dt (Check D22/D23 in XLSX)")
    if not mapped["_city_text"]:
        errors.append("Missing required fields: _city_text (City/Municipality)")

    status = "VALID" if not errors else "INVALID"
    return AforParsedRow(row_index=0, status=status, errors=errors, data=mapped)


GROUND_TRUTH = {
    "responder_type": "Augmenting Team",
    "fire_station_name": "Team Augment",
    "notification_dt_date": "2025-03-23",
    "region": "NCR",
    "province_district": "Metro Manila",
    "city_municipality": "Caloocan",
    "incident_address": "124, Brgy 179, Caloocan City",
    "nearest_landmark": "SM Farview",
    "caller_name": "Maria Isabel",
    "caller_number": "09727482631",
    "receiver_name": "FO1 Alice Guo",
    "engine_dispatched": "E232112-21, E232112-22",
    "time_engine_dispatched_hhmm": "14:31",
    "time_arrived_at_scene_hhmm": "14:38",
    "total_response_time_minutes": 69,
    "distance_km": 67,
    "alarm_level_normalized": "Second Alarm",
    "time_returned_hhmm": "15:48",
    "total_gas_consumed_liters": 67,
    "classification_of_involved": "Structural",
    "category": "Residential",
    "owner_name": "Maria Isabel",
    "general_description": "Concrete",
    "area_of_origin": "Kitchen / Cooking Area",
    "stage_of_fire": "Free-burning",
    "extent_of_damage": "Confined to Room",
    "extent_floor_sqm": 100.0,
    "extent_land_ha": 0.3,
    "structures_affected": 1,
    "households_affected": 2,
    "families_affected": 3,
    "individuals_affected": 4,
    "vehicles_affected": 5,
    "res_bfp_fire_trucks": 3,
    "res_lgu_owned_trucks": 1,
    "res_non_bfp_trucks": 40,
    "res_bfp_ambulance": 1,
    "res_non_bfp_ambulance": 42,
    "res_bfp_rescue_trucks": 43,
    "res_non_bfp_rescue_trucks": 44,
    "res_other_vehicles": 69,
    "tool_scba": 4,
    "tool_rope": 2,
    "tool_ladder": 1,
    "tool_hoseline": 3,
    "tool_hydrolic": 7,
    "tool_others": 2,
    "hydrant_distance": "250m",
    "alarm_foua_time": "14:38",
    "alarm_1st_time": "14:40",
    "alarm_2nd_time": "14:45",
    "alarm_3rd_time": "15:00",
    "alarm_4th_time": "16:00",
    "alarm_5th_time": "17:00",
    "alarm_tf_alpha_time": "18:00",
    "alarm_tf_bravo_time": "19:00",
    "alarm_tf_charlie_time": "20:00",
    "alarm_tf_delta_time": "21:00",
    "alarm_general_time": "22:00",
    "alarm_fuc_time": "15:45",
    "alarm_fo_time": "16:30",
    "alarm_foua_commander": "INSP Leo Villanueva (Ground Commander)",
    "injured_civilian_m": 2,
    "injured_civilian_f": 1,
    "injured_bfp_m": 1,
    "injured_bfp_f": 0,
    "pod_engine_commander": "SFO1 David Perez",
    "pod_shift_in_charge": "SFO2 Michael Torres",
    "pod_nozzleman": "FO1 John Ramos",
    "pod_lineman": "FO1 Carlos Dantes",
    "pod_engine_crew": "FO1 Kevin Lim, FO1 Albert Sy",
    "pod_dpo": "FO2 Robert Ocampo",
    "pod_safety_officer_name": "SFO1 Elena Cruz",
    "pod_safety_officer_contact": "0917-111-2222",
    "pod_fire_arson_investigator_name": "FO3 Sarah Gomez",
    "pod_fire_arson_investigator_contact": "0998-765-4321",
    "other_personnel_count": 9,
    "problems_count": 25,
    "icp_present": True,
    "icp_location_starts_with": "Specify location of ICP",
}


def to_hhmm(value: Any) -> str:
    if value is None:
        return ""
    s = str(value)
    m = re.search(r"(\d{2}:\d{2})", s)
    return m.group(1) if m else ""


def run_test(xlsx_path: str, ground_truth: dict | None = None):
    wb = load_workbook(xlsx_path, data_only=True)
    ws_name = next((n for n in wb.sheetnames if "AFOR" in n.upper()), wb.sheetnames[0])
    parser = BfpXlsxParser(wb[ws_name])
    raw = parser.parse()
    result = parse_afor_report_data(raw, region_id=1)

    ns = result.data.get("incident_nonsensitive_details", {})
    sens = result.data.get("incident_sensitive_details", {})
    rd = ns.get("resources_deployed", {})
    tl = ns.get("alarm_timeline", {})
    pod = sens.get("personnel_on_duty", {})
    caller_info = result.data.get("caller_info", "")
    c_name = caller_info.split("/")[0].strip() if "/" in caller_info else caller_info
    c_num = caller_info.split("/")[1].strip() if "/" in caller_info else ""

    failures = []
    passed = []

    def chk(label, actual, expected):
        a = str(actual or "").strip() if not isinstance(actual, (int, float, bool)) else actual
        e = str(expected or "").strip() if not isinstance(expected, (int, float, bool)) else expected
        if str(a).lower() == str(e).lower() or a == e:
            passed.append(f"  ✅ {label}: {repr(actual)}")
        else:
            failures.append(f"  ❌ {label}: got {repr(actual)!r} expected {repr(expected)!r}")

    chk("result.status", result.status, "VALID")

    if ground_truth:
        chk("responder_type", ns.get("responder_type"), ground_truth["responder_type"])
        chk("fire_station_name", ns.get("fire_station_name"), ground_truth["fire_station_name"])
        chk("notification_dt_date", str(ns.get("notification_dt") or "")[:10], ground_truth["notification_dt_date"])
        chk("region", ns.get("region") or result.data.get("_region_text"), ground_truth["region"])
        chk("province_district", ns.get("province_district") or result.data.get("_province_text"), ground_truth["province_district"])
        chk("city_municipality", ns.get("city_municipality") or result.data.get("_city_text"), ground_truth["city_municipality"])
        chk("incident_address", ns.get("incident_address") or sens.get("street_address"), ground_truth["incident_address"])
        chk("nearest_landmark", ns.get("nearest_landmark") or sens.get("landmark"), ground_truth["nearest_landmark"])
        chk("caller_name", c_name, ground_truth["caller_name"])
        chk("caller_number", c_num, ground_truth["caller_number"])
        chk("receiver_name", sens.get("receiver_name"), ground_truth["receiver_name"])
        chk("engine_dispatched", ns.get("engine_dispatched"), ground_truth["engine_dispatched"])
        chk("time_engine_dispatched_hhmm", to_hhmm(ns.get("time_engine_dispatched")), ground_truth["time_engine_dispatched_hhmm"])
        chk("time_arrived_at_scene_hhmm", to_hhmm(ns.get("time_arrived_at_scene")), ground_truth["time_arrived_at_scene_hhmm"])
        chk("total_response_time_minutes", ns.get("total_response_time_minutes"), ground_truth["total_response_time_minutes"])
        chk("distance_km", ns.get("distance_from_station_km") or ns.get("distance_to_fire_scene_km"), ground_truth["distance_km"])
        chk("alarm_level", ns.get("alarm_level"), ground_truth["alarm_level_normalized"])
        chk("time_returned_hhmm", to_hhmm(ns.get("time_returned_to_base")), ground_truth["time_returned_hhmm"])
        chk("total_gas_consumed_liters", ns.get("total_gas_consumed_liters"), ground_truth["total_gas_consumed_liters"])
        chk("classification_of_involved", ns.get("classification_of_involved") or ns.get("general_category"), ground_truth["classification_of_involved"])
        chk("category", ns.get("type_of_involved_general_category") or ns.get("sub_category"), ground_truth["category"])
        chk("owner_name", sens.get("owner_name"), ground_truth["owner_name"])
        chk("area_of_origin", ns.get("fire_origin"), ground_truth["area_of_origin"])
        chk("stage_of_fire", ns.get("stage_of_fire"), ground_truth["stage_of_fire"])
        chk("extent_of_damage", ns.get("extent_of_damage"), ground_truth["extent_of_damage"])
        chk("extent_floor_sqm", ns.get("extent_total_floor_area_sqm"), ground_truth["extent_floor_sqm"])
        chk("extent_land_ha", ns.get("extent_total_land_area_hectares"), ground_truth["extent_land_ha"])
        chk("structures_affected", ns.get("structures_affected"), ground_truth["structures_affected"])
        chk("households_affected", ns.get("households_affected"), ground_truth["households_affected"])
        chk("families_affected", ns.get("families_affected"), ground_truth["families_affected"])
        chk("individuals_affected", ns.get("individuals_affected"), ground_truth["individuals_affected"])
        chk("vehicles_affected", ns.get("vehicles_affected"), ground_truth["vehicles_affected"])
        chk("res_bfp_fire_trucks", rd.get("bfp_fire_trucks"), ground_truth["res_bfp_fire_trucks"])
        chk("res_lgu_owned_trucks", rd.get("lgu_owned_trucks"), ground_truth["res_lgu_owned_trucks"])
        chk("res_non_bfp_trucks", rd.get("non_bfp_trucks"), ground_truth["res_non_bfp_trucks"])
        chk("res_bfp_ambulance", rd.get("bfp_ambulance"), ground_truth["res_bfp_ambulance"])
        chk("res_non_bfp_ambulance", rd.get("non_bfp_ambulance"), ground_truth["res_non_bfp_ambulance"])
        chk("res_bfp_rescue_trucks", rd.get("bfp_rescue_trucks"), ground_truth["res_bfp_rescue_trucks"])
        chk("res_non_bfp_rescue_trucks", rd.get("non_bfp_rescue_trucks"), ground_truth["res_non_bfp_rescue_trucks"])
        chk("res_other_vehicles", rd.get("other_vehicles"), ground_truth["res_other_vehicles"])
        chk("tool_scba", rd.get("scba"), ground_truth["tool_scba"])
        chk("tool_rope", rd.get("rope"), ground_truth["tool_rope"])
        chk("tool_ladder", rd.get("ladder"), ground_truth["tool_ladder"])
        chk("tool_hoseline", rd.get("hoseline"), ground_truth["tool_hoseline"])
        chk("tool_hydrolic", rd.get("hydrolic_tools"), ground_truth["tool_hydrolic"])
        chk("tool_others", rd.get("other_tools"), ground_truth["tool_others"])
        chk("hydrant_distance", rd.get("hydrant_distance"), ground_truth["hydrant_distance"])

        chk("alarm_foua present", "alarm_foua" in tl or "foua" in tl, True)
        chk("alarm_foua_time", to_hhmm(tl.get("alarm_foua") or tl.get("foua")), ground_truth["alarm_foua_time"])
        chk("alarm_foua_commander", tl.get("alarm_foua_commander"), ground_truth["alarm_foua_commander"])
        chk("alarm_1st_time", to_hhmm(tl.get("alarm_1st")), ground_truth["alarm_1st_time"])
        chk("alarm_2nd_time", to_hhmm(tl.get("alarm_2nd")), ground_truth["alarm_2nd_time"])
        chk("alarm_3rd_time", to_hhmm(tl.get("alarm_3rd")), ground_truth["alarm_3rd_time"])
        chk("alarm_4th_time", to_hhmm(tl.get("alarm_4th")), ground_truth["alarm_4th_time"])
        chk("alarm_5th_time", to_hhmm(tl.get("alarm_5th")), ground_truth["alarm_5th_time"])
        chk("alarm_tf_alpha_time", to_hhmm(tl.get("alarm_tf_alpha")), ground_truth["alarm_tf_alpha_time"])
        chk("alarm_tf_bravo_time", to_hhmm(tl.get("alarm_tf_bravo")), ground_truth["alarm_tf_bravo_time"])
        chk("alarm_tf_charlie_time", to_hhmm(tl.get("alarm_tf_charlie")), ground_truth["alarm_tf_charlie_time"])
        chk("alarm_tf_delta_time", to_hhmm(tl.get("alarm_tf_delta")), ground_truth["alarm_tf_delta_time"])
        chk("alarm_general_time", to_hhmm(tl.get("alarm_general")), ground_truth["alarm_general_time"])
        chk("alarm_fuc_time", to_hhmm(tl.get("alarm_fuc")), ground_truth["alarm_fuc_time"])
        chk("alarm_fo_time", to_hhmm(tl.get("alarm_fo")), ground_truth["alarm_fo_time"])

        chk("pod_engine_commander", pod.get("engine_commander"), ground_truth["pod_engine_commander"])
        chk("pod_shift_in_charge", pod.get("shift_in_charge"), ground_truth["pod_shift_in_charge"])
        chk("pod_nozzleman", pod.get("nozzleman"), ground_truth["pod_nozzleman"])
        chk("pod_lineman", pod.get("lineman"), ground_truth["pod_lineman"])
        chk("pod_engine_crew", pod.get("engine_crew"), ground_truth["pod_engine_crew"])
        chk("pod_dpo", pod.get("driver"), ground_truth["pod_dpo"])
        so = pod.get("safety_officer", {})
        chk("safety_officer_name", so.get("name") if isinstance(so, dict) else so, ground_truth["pod_safety_officer_name"])
        chk("safety_officer_contact", so.get("contact") if isinstance(so, dict) else "", ground_truth["pod_safety_officer_contact"])
        inv = pod.get("fire_arson_investigator") or pod.get("investigator", {})
        chk("fire_arson_investigator_name", inv.get("name") if isinstance(inv, dict) else inv, ground_truth["pod_fire_arson_investigator_name"])
        chk("fire_arson_investigator_contact", inv.get("contact") if isinstance(inv, dict) else "", ground_truth["pod_fire_arson_investigator_contact"])
        chk("other_personnel_count", len(sens.get("other_personnel", [])), ground_truth["other_personnel_count"])
        chk("problems_count", len(ns.get("problems_encountered", [])), ground_truth["problems_count"])
        chk("icp_present", sens.get("is_icp_present"), ground_truth["icp_present"])
        chk("icp_location_starts_with", str(sens.get("icp_location") or "").startswith(ground_truth["icp_location_starts_with"]), True)

    print(f"\n{'='*60}")
    print(f"TEST: {xlsx_path}")
    print(f"{'='*60}")
    for p in passed:
        print(p)
    for f in failures:
        print(f)
    print(f"\n{'='*60}")
    print(f"PASSED: {len(passed)}  FAILED: {len(failures)}")
    if failures:
        print("⚠️  FAILURES FOUND — fix required")
    else:
        print("✅ ALL ASSERTIONS PASSED")
    return failures


def main() -> int:
    root = Path(__file__).resolve().parents[1]
    afors = root / "AFORs"
    files = [
        afors / "afor_filled.xlsx",
        afors / "afor_filled2.xlsx",
        afors / "afor_3.xlsx",
    ]

    missing = [str(p) for p in files if not p.exists()]
    if missing:
        print("Missing required AFOR file(s):")
        for m in missing:
            print(f"  - {m}")
        return 2

    all_failures = []
    all_failures.extend(run_test(str(files[0]), None))
    all_failures.extend(run_test(str(files[1]), None))
    all_failures.extend(run_test(str(files[2]), GROUND_TRUTH))

    if all_failures:
        print("\nOverall result: FAIL")
        return 1

    print("\nOverall result: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
