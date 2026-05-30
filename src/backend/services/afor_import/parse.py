"""AFOR workbook/CSV parsing — template detection, row mapping, validation."""

from __future__ import annotations

import csv
import io
import re
from datetime import datetime, timedelta
from typing import Any

from services.afor_import.models import (
    AforFormKind,
    AforParsedRow,
    AforParseResponse,
    WildlandRowSource,
)

ALARM_LEVEL_MAP = {
    "1ST": "First Alarm",
    "1ST ALARM": "First Alarm",
    "FIRST": "First Alarm",
    "FIRST ALARM": "First Alarm",
    "2ND": "Second Alarm",
    "2ND ALARM": "Second Alarm",
    "SECOND": "Second Alarm",
    "SECOND ALARM": "Second Alarm",
    "3RD": "Third Alarm",
    "3RD ALARM": "Third Alarm",
    "THIRD": "Third Alarm",
    "THIRD ALARM": "Third Alarm",
    "4TH": "Fourth Alarm",
    "4TH ALARM": "Fourth Alarm",
    "FOURTH": "Fourth Alarm",
    "FOURTH ALARM": "Fourth Alarm",
    "5TH": "Fifth Alarm",
    "5TH ALARM": "Fifth Alarm",
    "FIFTH": "Fifth Alarm",
    "FIFTH ALARM": "Fifth Alarm",
    "TF ALPHA": "Task Force Alpha",
    "TASK FORCE ALPHA": "Task Force Alpha",
    "TF BRAVO": "Task Force Bravo",
    "TASK FORCE BRAVO": "Task Force Bravo",
    "TF CHARLIE": "Task Force Charlie",
    "TASK FORCE CHARLIE": "Task Force Charlie",
    "TF DELTA": "Task Force Delta",
    "TASK FORCE DELTA": "Task Force Delta",
    "GENERAL": "General Alarm",
    "GENERAL ALARM": "General Alarm",
}

# ---------------------------------------------------------------------------
# Safe value conversions (shared by structural and wildland parsers)
# ---------------------------------------------------------------------------


def _safe_int(val: Any, default: int = 0) -> int:
    if val is None or val == "" or val == "N/A":
        return default
    try:
        if isinstance(val, (int, float)):
            return int(val)
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return default


def _safe_float(val: Any, default: float = 0.0) -> float:
    if val is None or val == "" or val == "N/A":
        return default
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return default


def _safe_dt(val: Any) -> str | None:
    """Safe datetime string conversion."""
    if isinstance(val, datetime):
        return val.isoformat()
    if not val:
        return None

    # Excel stores dates/times as serial floats in many filled templates.
    # Serial date epoch (Windows): 1899-12-30.
    if isinstance(val, (int, float)):
        try:
            serial = float(val)
            base = datetime(1899, 12, 30)
            dt = base + timedelta(days=serial)
            if serial < 1:
                return dt.strftime("%H:%M:%S")
            return dt.isoformat()
        except Exception:
            return None

    raw_numeric = str(val).strip()
    try:
        serial = float(raw_numeric)
        base = datetime(1899, 12, 30)
        dt = base + timedelta(days=serial)
        if serial < 1:
            return dt.strftime("%H:%M:%S")
        return dt.isoformat()
    except (ValueError, TypeError):
        pass

    dt_str = str(val).strip()
    for fmt in [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%m-%d-%Y %H:%M:%S",
        "%m-%d-%Y %H:%M",
        "%H:%M",
        "%H:%M:%S",
        "%Y-%m-%d",
        "%m-%d-%Y",
        "%m/%d/%Y",
    ]:
        try:
            return datetime.strptime(dt_str, fmt).isoformat()
        except ValueError:
            continue
    return None


_COORD_RE = re.compile(r"^([A-Z]+)(\d+)$")


class _SheetCell:
    def __init__(self, value: Any):
        self.value = value


def _column_letters_to_index(letters: str) -> int:
    index = 0
    for char in letters:
        index = (index * 26) + (ord(char) - ord("A") + 1)
    return index - 1


# ---------------------------------------------------------------------------
# CSV worksheet adapter (exposes CSV rows through A1-cell notation)
# ---------------------------------------------------------------------------


class CsvWorksheetAdapter:
    """Expose CSV cells through worksheet-like `A1` coordinates."""

    def __init__(self, rows: list[list[str]]):
        self.rows = rows

    def __getitem__(self, coord: str) -> _SheetCell:
        match = _COORD_RE.match(coord.upper())
        if not match:
            raise KeyError(f"Invalid coordinate: {coord}")

        column_letters, row_number = match.groups()
        row_idx = int(row_number) - 1
        col_idx = _column_letters_to_index(column_letters)

        value = None
        if 0 <= row_idx < len(self.rows) and 0 <= col_idx < len(self.rows[row_idx]):
            raw_value = self.rows[row_idx][col_idx]
            if isinstance(raw_value, str):
                raw_value = raw_value.strip()
            value = raw_value or None

        return _SheetCell(value)


def _looks_like_official_afor_csv(rows: list[list[str]]) -> bool:
    if not rows:
        return False

    first_column_values = [
        (row[0].strip().upper() if row and isinstance(row[0], str) else "") for row in rows
    ]
    return (
        "AFTER FIRE OPERATIONS REPORT" in first_column_values
        and "A. RESPONSE DETAILS" in first_column_values
    )


# ---------------------------------------------------------------------------
# Structural sheet helpers
# ---------------------------------------------------------------------------


def _cell_str(ws: Any, coord: str) -> str:
    try:
        v = ws[coord].value
    except Exception:
        return ""
    if v is None:
        return ""
    return str(v).strip()


def _sheet_has_structural_markers(ws: Any) -> bool:
    """Structural AFOR marker detection, tolerant to row shifts in filled templates."""
    title_row, section_row = _find_structural_marker_rows(ws)
    if title_row is None or section_row is None:
        return False
    # In official templates, section header appears a few rows after title.
    return 2 <= (section_row - title_row) <= 8


def _find_structural_marker_rows(ws: Any) -> tuple[int | None, int | None]:
    """Find title/section marker rows by scanning the top-left block of the sheet."""
    title_row: int | None = None
    section_row: int | None = None

    for row in range(1, 161):
        row_values = [
            _cell_str(ws, f"{col}{row}").upper() for col in ("A", "B", "C", "D", "E", "F")
        ]
        combined = " ".join(v for v in row_values if v).strip()

        if title_row is None and "AFTER FIRE OPERATIONS REPORT" in combined:
            title_row = row
        if section_row is None and "A. RESPONSE DETAILS" in combined:
            section_row = row

        if title_row is not None and section_row is not None:
            break

    return title_row, section_row


def _sheet_has_wildland_markers(ws: Any) -> bool:
    """
    Wildland workbook: main sheet title in B12 and section A header in B13.
    Tie-break: sheet name containing 'WILDLAND FIRE AFOR' wins over structural when both match.
    """
    b12 = _cell_str(ws, "B12").upper()
    b13 = _cell_str(ws, "B13").upper()
    if "WILDLAND" in b12 and "A. DATES" in b13:
        return True
    if "WILDLAND FIRE" in b12:
        return True
    return False


# ---------------------------------------------------------------------------
# Template detection
# ---------------------------------------------------------------------------


def detect_afor_template_kind(wb: Any) -> AforFormKind | None:
    """
    Classify uploaded workbook as structural vs wildland AFOR.

    Detection:
        Wildland markers also match when B12 contains WILDLAND FIRE even if B13 is not the usual
        "A. DATES…" line (see `_sheet_has_wildland_markers`).

    Rules (order):
    1. If any sheet name contains 'WILDLAND FIRE AFOR' (case-insensitive) and that sheet
       has wildland markers (B12/B13 or title containing WILDLAND) → WILDLAND_AFOR.
    2. Else if any sheet has structural markers (A14 + A18) → STRUCTURAL_AFOR.
    3. Else if any sheet has wildland markers without relying on sheet name → WILDLAND_AFOR.
    4. Else None (ambiguous).
    """
    sheets: list[tuple[str, Any]] = [(n, wb[n]) for n in wb.sheetnames]

    for name, ws in sheets:
        if "WILDLAND FIRE AFOR" in name.upper() and _sheet_has_wildland_markers(ws):
            return "WILDLAND_AFOR"

    for name, ws in sheets:
        if _sheet_has_structural_markers(ws):
            return "STRUCTURAL_AFOR"

    for _name, ws in sheets:
        if _sheet_has_wildland_markers(ws):
            return "WILDLAND_AFOR"

    return None


def _pick_structural_worksheet(wb: Any) -> Any:
    for name in wb.sheetnames:
        ws = wb[name]
        if _sheet_has_structural_markers(ws):
            return ws
    for name in wb.sheetnames:
        if "AFOR" in name.upper():
            return wb[name]
    return wb.active


# ---------------------------------------------------------------------------
# Wildland helpers
# ---------------------------------------------------------------------------


def _pick_wildland_worksheet(wb: Any) -> Any:
    for name in wb.sheetnames:
        if "WILDLAND" in name.upper() and "AFOR" in name.upper():
            return wb[name]
    for name in wb.sheetnames:
        if _sheet_has_wildland_markers(wb[name]):
            return wb[name]
    return wb.active


_WILDLAND_FIRE_TYPES_LOWER = {
    "fire",
    "agricultural land fire",
    "brush fire",
    "forest fire",
    "grassland fire",
    "grazing land fire",
    "mineral land fire",
    "peatland fire",
}


def _normalize_wildland_fire_type(raw: Any) -> str | None:
    if raw is None:
        return None
    t = str(raw).strip().lower()
    if t in _WILDLAND_FIRE_TYPES_LOWER:
        return t
    return None


def _parse_ha_from_area_text(raw: Any) -> float | None:
    if raw is None:
        return None
    s = str(raw).strip()
    m = re.search(r"([\d.]+)\s*ha", s, re.IGNORECASE)
    if m:
        try:
            return float(m.group(1))
        except ValueError:
            return None
    return None


# ---------------------------------------------------------------------------
# Wildland parser
# ---------------------------------------------------------------------------


class WildlandXlsxParser:
    """Parser for BFP wildland AFOR workbook (sheet 'WILDLAND FIRE AFOR')."""

    def __init__(self, ws: Any):
        self.ws = ws

    def get(self, coord: str) -> Any:
        val = self.ws[coord].value
        if val is None:
            return None
        if isinstance(val, str):
            return val.strip()
        return val

    def parse(self) -> dict[str, Any]:
        def _dt_cell(coord: str) -> datetime | None:
            v = self.get(coord)
            if isinstance(v, datetime):
                return v
            return None

        call_received = _dt_cell("D15")
        fire_started = _dt_cell("D17")
        fire_arrival = _dt_cell("D19")
        fire_controlled = _dt_cell("D21")

        extras: list[str] = []
        for coord in ("E28", "E29"):
            v = self.get(coord)
            if v:
                extras.append(str(v))

        fire_behavior = {
            "elevation_ft": _safe_float(self.get("D51"), 0.0) or None,
            "relative_position_slope": self.get("D52"),
            "aspect": self.get("D53"),
            "flame_length_ft": _safe_float(self.get("D54"), 0.0) or None,
            "rate_of_spread_chains_per_hour": _safe_float(self.get("D55"), 0.0) or None,
        }

        problems: list[str] = []
        for r in range(76, 80):
            line = self.get(f"B{r}")
            if line and str(line).strip():
                problems.append(str(line).strip())

        recommendations: list[str] = []
        for r in range(83, 87):
            line = self.get(f"B{r}")
            if line and str(line).strip():
                recommendations.append(str(line).strip())

        alarm_rows: list[dict[str, Any]] = []
        for r in range(50, 65):
            status = self.get(f"J{r}")
            if not status or not str(status).strip():
                continue
            time_declared = self.get(f"K{r}")
            commander = self.get(f"L{r}")
            alarm_rows.append(
                {
                    "alarm_status": str(status).strip(),
                    "time_declared": str(time_declared).strip() if time_declared else "",
                    "ground_commander": str(commander).strip() if commander else "",
                }
            )

        raw_type = self.get("G44")
        wft = _normalize_wildland_fire_type(raw_type)

        return {
            "call_received_at": call_received,
            "fire_started_at": fire_started,
            "fire_arrival_at": fire_arrival,
            "fire_controlled_at": fire_controlled,
            "caller_transmitted_by": self.get("B33") or self.get("B32"),
            "caller_office_address": self.get("D33") or self.get("D32"),
            "call_received_by_personnel": self.get("F33") or self.get("F32"),
            "engine_dispatched": self.get("D23"),
            "incident_location_description": self.get("D31") or self.get("B31"),
            "distance_to_fire_station_km": _safe_float(self.get("D32"), 0.0)
            if self.get("D32") not in (None, "")
            else None,
            "primary_action_taken": self.get("E27"),
            "assistance_combined_summary": " | ".join(extras) if extras else None,
            "buildings_involved": _safe_int(self.get("B40")),
            "buildings_threatened": _safe_int(self.get("G40")),
            "ownership_and_property_notes": self.get("B41") or self.get("B39"),
            "total_area_burned_display": self.get("B44"),
            "total_area_burned_hectares": _parse_ha_from_area_text(self.get("B44")),
            "wildland_fire_type": wft,
            "raw_wildland_fire_type": raw_type,
            "area_type_summary": {},
            "causes_and_ignition_factors": {},
            "suppression_factors": {},
            "weather": {},
            "fire_behavior": {k: v for k, v in fire_behavior.items() if v not in (None, "", 0.0)},
            "peso_losses": {},
            "casualties": {},
            "narration": self.get("B68"),
            "problems_encountered": problems,
            "recommendations_list": recommendations,
            "prepared_by": self.get("B91"),
            "prepared_by_title": self.get("B92"),
            "noted_by": self.get("E88") or self.get("F88"),
            "noted_by_title": self.get("E91"),
            "wildland_alarm_statuses": alarm_rows,
            "wildland_assistance_rows": [],
        }


def parse_wildland_afor_report_data(data: dict[str, Any], region_id: int) -> AforParsedRow:
    """Map wildland workbook dict into commit payload + validation."""
    errors: list[str] = []

    primary = (data.get("primary_action_taken") or "").strip()
    engine = (data.get("engine_dispatched") or "").strip()
    narration = (data.get("narration") or "").strip()
    call_at = data.get("call_received_at")
    wft = data.get("wildland_fire_type")

    if not primary and not engine and not narration and not call_at and not wft:
        errors.append(
            "Missing wildland content: need at least one of call time (D15), primary action (E27), "
            "engine (D23), narration (B68), or wildland fire type (G44)."
        )

    if data.get("raw_wildland_fire_type") and not wft:
        errors.append(
            f"Wildland fire type value is not allowed: {data.get('raw_wildland_fire_type')!r}. "
            "Use the Sheet1 list (e.g. Brush Fire, Forest Fire)."
        )

    wl_payload = {
        k: v for k, v in data.items() if k not in ("raw_wildland_fire_type", "recommendations_list")
    }
    wl_payload["recommendations"] = data.get("recommendations_list") or []

    mapped: dict[str, Any] = {
        "_form_kind": "WILDLAND_AFOR",
        "_city_text": "",
        "region_id": region_id,
        "wildland": wl_payload,
    }

    status = "VALID" if not errors else "INVALID"
    return AforParsedRow(row_index=0, status=status, errors=errors, data=mapped)


# ---------------------------------------------------------------------------
# Structural row mapper
# ---------------------------------------------------------------------------


def _combine_date_and_time(notification_dt: str | None, time_value: Any) -> str | None:
    if not notification_dt or not time_value:
        return None

    date_part = str(notification_dt).split("T", 1)[0]
    return _safe_dt(f"{date_part} {str(time_value).strip()}")


def _time_str(val: Any) -> str | None:
    """Extract HH:MM 24-hour string from a raw time value (datetime.time, Excel serial float, or string)."""
    if val is None:
        return None
    # Native Python datetime.time or datetime object
    if hasattr(val, "hour") and hasattr(val, "minute"):
        return f"{val.hour:02d}:{val.minute:02d}"
    try:
        t_serial = float(val)
        base = datetime(1899, 12, 30)
        time_dt = (base + timedelta(days=t_serial)).time()
        return f"{time_dt.hour:02d}:{time_dt.minute:02d}"
    except (TypeError, ValueError):
        pass
    s = str(val).strip()
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return s or None


def _extract_barangay_from_address(address: str) -> str:
    """Extract barangay from AFOR D27 address.
    Tries keyword detection first (handles free-form input),
    then falls back to the AFOR template position (index 2 of comma-split)."""
    if not address or address.startswith("("):
        return ""
    # Keyword-based: match "Brgy.", "Bgy.", "Barangay " followed by the name until the next comma
    m = re.search(r"((?:Brgy|Bgy)\.?\s*[^,]+|Barangay\s+[^,]+)", address, re.IGNORECASE)
    if m:
        return m.group(1).strip().rstrip(".")
    # Positional fallback: AFOR template = "HouseNo, Street, Barangay, City, Province"
    parts = [p.strip() for p in address.split(",")]
    return parts[2] if len(parts) >= 3 else ""


class BfpXlsxParser:
    """Parser for the official BFP manual entry form (AFOR)."""

    def __init__(self, ws):
        self.ws = ws
        self._row_offset = self._infer_row_offset()

    def _infer_row_offset(self) -> int:
        """Infer row offset when users fill a structurally identical AFOR with shifted rows."""
        title_row, section_row = _find_structural_marker_rows(self.ws)
        if title_row is None:
            return 0

        offset = title_row - 14
        # Validate offset with section marker when available.
        if section_row is not None and (section_row - 18) != offset:
            return 0
        return offset

    def _coord_with_offset(self, coord: str) -> str:
        match = _COORD_RE.match(coord.upper())
        if not match or self._row_offset == 0:
            return coord

        col, row_str = match.groups()
        shifted_row = max(1, int(row_str) + self._row_offset)
        return f"{col}{shifted_row}"

    def get(self, coord: str) -> Any:
        shifted_coord = self._coord_with_offset(coord)
        val = self.ws[shifted_coord].value
        if val is None and shifted_coord != coord:
            # Fallback to canonical location to support mixed/custom sheets.
            val = self.ws[coord].value
        if val is None:
            return None
        if isinstance(val, str):
            return val.strip()
        return val

    def _is_marked(self, coord: str) -> bool:
        raw = self.get(coord)
        if raw is None:
            return False

        if isinstance(raw, bool):
            return raw

        if isinstance(raw, (int, float)):
            return raw != 0

        val = str(raw).strip().lower()
        if not val:
            return False

        if val.startswith("="):
            expr = val.lstrip("=").strip().lower()
            if expr in {"true", "1"}:
                return True

        return val in {
            "x",
            "1",
            "true",
            "v",
            "/",
            "yes",
            "checked",
            "☑",
            "☒",
            "✓",
            "✔",
            "✅",
        }

    def _first_nonempty(self, *coords: str) -> Any:
        for coord in coords:
            val = self.get(coord)
            if val is None:
                continue
            if isinstance(val, str) and not val.strip():
                continue
            return val
        return None

    def _male_female_pair(self, row: int) -> tuple[Any, Any]:
        # Some AFOR variants shift M/F columns by one; try common adjacent pairs.
        candidate_pairs = [("D", "E"), ("C", "D"), ("E", "F"), ("F", "G")]
        fallback_pair = (None, None)
        for male_col, female_col in candidate_pairs:
            male_val = self.get(f"{male_col}{row}")
            female_val = self.get(f"{female_col}{row}")
            if fallback_pair == (None, None):
                fallback_pair = (male_val, female_val)

            has_male = male_val not in (None, "")
            has_female = female_val not in (None, "")
            if has_male or has_female:
                return male_val, female_val

        return fallback_pair

    def _is_marked_on_row(self, row: int, cols: tuple[str, ...] = ("B", "C", "D")) -> bool:
        return any(self._is_marked(f"{col}{row}") for col in cols)

    def _parse_name_contact(self, coord: str, prefix: str) -> dict[str, Any]:
        """Read a cell and split on '/' to get name and contact number."""
        raw = str(self.get(coord) or "").strip()
        if "/" in raw:
            parts = raw.split("/", 1)
            return {prefix: parts[0].strip(), f"{prefix}_contact": parts[1].strip()}
        return {prefix: raw, f"{prefix}_contact": ""}

    def parse(self) -> dict[str, Any]:
        """Extract sections A through L into a comprehensive data dictionary."""

        # Section A: Response Details
        responder_type = (
            "First Responder"
            if self._is_marked("B20")
            else ("Augmenting Team" if self._is_marked("B21") else "First Responder")
        )

        # Section B: Classification
        classification = "Structural"
        cat_val = self.get("D48")
        if self._is_marked_on_row(49):
            classification = "Non-Structural"
            cat_val = self.get("D49")
        elif self._is_marked_on_row(50):
            classification = "Transportation"
            cat_val = self.get("D50")
        elif self.get("D49") not in (None, ""):
            classification = "Non-Structural"
            cat_val = self.get("D49")
        elif self.get("D50") not in (None, ""):
            classification = "Transportation"
            cat_val = self.get("D50")

        stage = self.get("D54") or self.get("B54")
        if stage and "pick from dropdown" in str(stage).lower():
            stage = None

        # Extent of Damage
        extent = "None / Minor"
        if self._is_marked_on_row(57):
            extent = "Confined to Object"
        elif self._is_marked_on_row(58):
            extent = "Confined to Room"
        elif self._is_marked_on_row(59):
            extent = "Confined to Structure"
        elif self._is_marked_on_row(60):
            extent = "Total Loss"
        elif self._is_marked_on_row(61):
            extent = "Extended Beyond Structure"
        else:
            extent_text = str(self._first_nonempty("D57", "D58", "D59", "D60", "D61") or "").strip()
            if extent_text:
                extent = extent_text

        # Section J: Problems
        problems = []
        prob_map = {
            "B195": "Inaccurate address",
            "B196": "Geographically challenged",
            "B197": "Road conditions",
            "B198": "Road under construction",
            "B199": "Traffic congestion",
            "B200": "Road accidents",
            "B201": "Vehicles failure to yield",
            "B202": "Natural Disasters",
            "B203": "Civil Disturbance",
            "B204": "Uncooperative or panicked residents",
            "B205": "Safety and security threats",
            "B206": "Property security or owner delays",
            "B207": "Engine failure",
            "B208": "Uncooperative fire auxiliary",
            "B209": "Poor water supply access",
            "B210": "Intense heat and smoke",
            "B211": "Structural hazards",
            "B212": "Equipment malfunction",
            "B213": "Poor inter-agency coordination",
            "B214": "Radio communication breakdown",
            "B215": "HazMat risks",
            "B216": "Physical exhaustion and injuries",
            "B217": "Emotional and psychological effects",
            "B218": "Community complaints",
            "B219": "Others",
        }
        for c, flavor in prob_map.items():
            row_num = int(c[1:])
            if c == "B219":
                others_text = (str(self.get("C219") or "")).strip()
                if self._is_marked_on_row(row_num) or others_text:
                    label = others_text if others_text else "Others"
                    problems.append(label)
            elif self._is_marked_on_row(row_num):
                problems.append(flavor)

        icp_present = self._is_marked_on_row(102)
        _raw_icp = str(self.get("D102") or "").strip()
        _raw_icp = re.sub(r"(?i)^Specify\s+location\s+of\s+ICP\s*:\s*", "", _raw_icp).strip()
        icp_location = _raw_icp or None if icp_present else None

        # Section I: Narrative joining (Rows 160 to 190)
        narrative_lines = []
        for r in range(160, 191):
            line = self.get(f"B{r}")
            if line:
                narrative_lines.append(str(line))

        # Section G: Other Personnel (Rows 124 to 132)
        others = []
        for r in range(124, 133):
            name = self.get(f"B{r}")
            rem = self.get(f"E{r}")
            if name and "N/A" not in str(name).upper():
                others.append({"name": name, "designation": rem or ""})

        inj_civ_m, inj_civ_f = self._male_female_pair(106)
        inj_bfp_m, inj_bfp_f = self._male_female_pair(107)
        inj_aux_m, inj_aux_f = self._male_female_pair(108)
        fat_civ_m, fat_civ_f = self._male_female_pair(109)
        fat_bfp_m, fat_bfp_f = self._male_female_pair(110)
        fat_aux_m, fat_aux_f = self._male_female_pair(111)

        return {
            "responder_type": responder_type,
            "fire_station_name": self.get("D20")
            if responder_type == "First Responder"
            else self.get("D21"),
            "notification_date": self.get("D22"),
            "notification_time": self.get("D23"),
            "region": self.get("D24"),
            "province": self.get("D25"),
            "city": self.get("D26"),
            "address": self.get("D27"),
            "barangay": _extract_barangay_from_address(self.get("D27") or ""),
            "landmark": self.get("D28"),
            "caller_info": self.get("D29"),
            "receiver": self.get("D30"),
            "engine": self.get("D31"),
            "engine_2": self.get("D32"),
            "engine_3": self.get("D33"),
            "time_dispatched": self.get("D34"),
            "time_dispatched_2": self.get("D35"),
            "time_dispatched_3": self.get("D36"),
            "time_arrived": self.get("D37"),
            "time_arrived_2": self.get("D38"),
            "time_arrived_3": self.get("D39"),
            "response_time": self.get("D40"),
            "distance_km": self.get("D41"),
            "alarm_level": self.get("D42"),
            "time_returned": self.get("D43"),
            "gas_liters": self.get("D44"),
            "classification": classification,
            "category": cat_val,
            "owner": self.get("D51"),
            "description": self.get("D52"),
            "origin": self.get("D53"),
            "stage": stage,
            "extent": extent,
            "extent_total_floor_area_sqm": (
                self.get("D58")
                if extent == "Confined to Room"
                else self.get("D59")
                if extent == "Confined to Structure"
                else self.get("D60")
                if extent in ("Total Loss", "Extended Beyond Structure")
                else None
            ),
            "extent_total_land_area_hectares": (
                self.get("E59") or self.get("D60")
                if extent == "Confined to Structure"
                else self.get("E60")
                if extent == "Total Loss"
                else None
            ),
            "extent_description": (
                str(self.get("D56") or "").strip() or None
                if extent == "None / Minor"
                else str(self.get("D57") or "").strip() or None
                if extent == "Confined to Object"
                else str(self.get("D61") or "").strip() or None
                if extent == "Extended Beyond Structure"
                else None
            ),
            "extent_objects_count": (
                _safe_int(self.get("D61")) if extent == "Extended Beyond Structure" else None
            ),
            "struct_aff": self.get("D62"),
            "house_aff": self.get("D63"),
            "fam_aff": self.get("D64"),
            "indiv_aff": self.get("D65"),
            "vehic_aff": self.get("D66"),
            "res_bfp_truck": self.get("D70"),
            "res_lgu_truck": self.get("D71"),
            "res_vol_truck": self.get("D72"),
            "res_bfp_amb": self.get("D73"),
            "res_non_amb": self.get("D74"),
            "res_bfp_resc": self.get("D75"),
            "res_non_resc": self.get("D76"),
            "res_others": self.get("D77"),
            "tool_scba": self.get("D79"),
            "tool_rope": self.get("D80"),
            "tool_ladder": self.get("D81"),
            "tool_hose": self.get("D82"),
            "tool_hydra": self.get("D83"),
            "tool_others": self.get("D84"),
            "hydrant_dist": self.get("D85"),
            "timeline": {
                "alarm_foua": {
                    "time": self.get("D88"),
                    "date": self.get("E88"),
                    "commander": self.get("F88"),
                },
                "alarm_1st": {
                    "time": self.get("D89"),
                    "date": self.get("E89"),
                    "commander": self.get("F89"),
                },
                "alarm_2nd": {
                    "time": self.get("D90"),
                    "date": self.get("E90"),
                    "commander": self.get("F90"),
                },
                "alarm_3rd": {
                    "time": self.get("D91"),
                    "date": self.get("E91"),
                    "commander": self.get("F91"),
                },
                "alarm_4th": {
                    "time": self.get("D92"),
                    "date": self.get("E92"),
                    "commander": self.get("F92"),
                },
                "alarm_5th": {
                    "time": self.get("D93"),
                    "date": self.get("E93"),
                    "commander": self.get("F93"),
                },
                "tf_alpha": {
                    "time": self.get("D94"),
                    "date": self.get("E94"),
                    "commander": self.get("F94"),
                },
                "tf_bravo": {
                    "time": self.get("D95"),
                    "date": self.get("E95"),
                    "commander": self.get("F95"),
                },
                "tf_charlie": {
                    "time": self.get("D96"),
                    "date": self.get("E96"),
                    "commander": self.get("F96"),
                },
                "tf_delta": {
                    "time": self.get("D97"),
                    "date": self.get("E97"),
                    "commander": self.get("F97"),
                },
                "general": {
                    "time": self.get("D98"),
                    "date": self.get("E98"),
                    "commander": self.get("F98"),
                },
                "fuc": {
                    "time": self.get("D99"),
                    "date": self.get("E99"),
                    "commander": self.get("F99"),
                },
                "fo": {
                    "time": self.get("D100"),
                    "date": self.get("E100"),
                    "commander": self.get("F100"),
                },
            },
            "icp_present": icp_present,
            "icp_location": icp_location,
            "inj_civ_m": inj_civ_m,
            "inj_civ_f": inj_civ_f,
            "inj_bfp_m": inj_bfp_m,
            "inj_bfp_f": inj_bfp_f,
            "inj_aux_m": inj_aux_m,
            "inj_aux_f": inj_aux_f,
            "fat_civ_m": fat_civ_m,
            "fat_civ_f": fat_civ_f,
            "fat_bfp_m": fat_bfp_m,
            "fat_bfp_f": fat_bfp_f,
            "fat_aux_m": fat_aux_m,
            "fat_aux_f": fat_aux_f,
            "pod_commander": self.get("D114"),
            "pod_shift": self.get("D115"),
            "pod_nozzleman": self.get("D116"),
            "pod_lineman": self.get("D117"),
            "pod_crew": self.get("D118"),
            "pod_dpo": self.get("D119"),
            **self._parse_name_contact("D120", "pod_safety"),
            **self._parse_name_contact("D121", "pod_inv"),
            "others_list": others,
            "narrative": "\n".join(narrative_lines),
            "problems": problems,
            "recommendations": self.get("B222"),
            "disposition": self.get("B229"),
            "prepared_by": self._first_nonempty("C239", "C240", "C238"),
            "noted_by": self._first_nonempty("E239", "E240", "F238"),
            # Backward-compatible aliases used by older tests/scripts.
            "extent_of_damage": extent,
            "structures_affected": self.get("D62"),
            "res_bfp_trucks": self.get("D70"),
            "alarm_1st": self.get("D89"),
        }


# ---------------------------------------------------------------------------
# Row mappers
# ---------------------------------------------------------------------------


def parse_afor_report_data(data: dict, region_id: int) -> AforParsedRow:
    """Map the extracted AFOR dictionary into the strict database schema."""
    errors: list[str] = []

    def _dt(d: Any, t: Any = None) -> str | None:
        if not d:
            return None

        if t:
            # Native Excel conversions often give datetime/date + datetime.time objects.
            if isinstance(d, datetime) and hasattr(t, "hour") and hasattr(t, "minute"):
                try:
                    return datetime.combine(d.date(), t).isoformat()
                except Exception:
                    pass

            # Excel serial date/time support for real filled XLSX exports.
            d_serial: float | None = None
            t_serial: float | None = None
            try:
                d_serial = float(d)
                t_serial = float(t)
            except (TypeError, ValueError):
                d_serial = None
                t_serial = None

            if d_serial is not None and t_serial is not None:
                try:
                    base = datetime(1899, 12, 30)
                    date_dt = base + timedelta(days=d_serial)
                    time_dt = base + timedelta(days=t_serial)
                    merged = datetime.combine(date_dt.date(), time_dt.time())
                    return merged.isoformat()
                except Exception:
                    pass

            date_part = d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d).split(" ")[0]
            return _safe_dt(f"{date_part} {str(t).strip()}")

        return _safe_dt(d)

    notif_dt = _dt(data.get("notification_date"), data.get("notification_time"))

    # Split caller_info: "Name / 0917-..."
    ci = str(data.get("caller_info") or "")
    c_name = ci.split("/")[0].strip() if "/" in ci else ci
    c_num = ci.split("/")[1].strip() if "/" in ci else ""

    casualty_details = {
        "injured": {
            "civilian": {
                "m": _safe_int(data.get("inj_civ_m")),
                "f": _safe_int(data.get("inj_civ_f")),
            },
            "firefighter": {
                "m": _safe_int(data.get("inj_bfp_m")),
                "f": _safe_int(data.get("inj_bfp_f")),
            },
            "auxiliary": {
                "m": _safe_int(data.get("inj_aux_m")),
                "f": _safe_int(data.get("inj_aux_f")),
            },
        },
        "fatalities": {
            "civilian": {
                "m": _safe_int(data.get("fat_civ_m")),
                "f": _safe_int(data.get("fat_civ_f")),
            },
            "firefighter": {
                "m": _safe_int(data.get("fat_bfp_m")),
                "f": _safe_int(data.get("fat_bfp_f")),
            },
            "auxiliary": {
                "m": _safe_int(data.get("fat_aux_m")),
                "f": _safe_int(data.get("fat_aux_f")),
            },
        },
    }

    timeline = data.get("timeline") or {
        "alarm_foua": {"time": None, "date": data.get("notification_date")},
        "alarm_1st": {
            "time": data.get("alarm_1st"),
            "date": data.get("notification_date"),
        },
        "alarm_2nd": {"time": None, "date": data.get("notification_date")},
        "alarm_3rd": {"time": None, "date": data.get("notification_date")},
        "alarm_4th": {"time": None, "date": data.get("notification_date")},
        "alarm_5th": {"time": None, "date": data.get("notification_date")},
        "tf_alpha": {"time": None, "date": data.get("notification_date")},
        "tf_bravo": {"time": None, "date": data.get("notification_date")},
        "tf_charlie": {"time": None, "date": data.get("notification_date")},
        "tf_delta": {"time": None, "date": data.get("notification_date")},
        "general": {"time": None, "date": data.get("notification_date")},
        "fuc": {"time": None, "date": data.get("notification_date")},
        "fo": {"time": None, "date": data.get("notification_date")},
    }

    # Normalize extent values to match the frontend form radio values
    _EXTENT_NORMALIZE: dict[str, str] = {
        "None / Minor": "None / Minor Damage",
        "Confined to Object": "Confined to Object/Vehicle",
        "Confined to Structure": "Confined to Structure or Property",
        "Extended Beyond Structure": "Extended Beyond Structure or Property",
    }
    raw_extent = data.get("extent") or data.get("extent_of_damage")
    normalized_extent = _EXTENT_NORMALIZE.get(str(raw_extent or "").strip(), raw_extent)

    # Build multi-engine list for alarm_timeline._engines
    engines: list[dict[str, Any]] = []
    for name_key, disp_key, arr_key in [
        ("engine", "time_dispatched", "time_arrived"),
        ("engine_2", "time_dispatched_2", "time_arrived_2"),
        ("engine_3", "time_dispatched_3", "time_arrived_3"),
    ]:
        eng_name = data.get(name_key)
        if eng_name:
            engines.append(
                {
                    "name": str(eng_name).strip(),
                    "time_dispatched": _time_str(data.get(disp_key)),
                    "time_arrived": _time_str(data.get(arr_key)),
                }
            )

    def _alarm_entry(key: str) -> dict | None:
        t = timeline.get(key) or {}
        if not isinstance(t, dict):
            return None
        dt_val = _dt(t.get("date"), t.get("time"))
        cmd = (str(t.get("commander") or "")).strip() or None
        if dt_val or cmd:
            return {"time": dt_val, "commander": cmd}
        return None

    incident_nonsensitive_details = {
        "notification_dt": notif_dt,
        "responder_type": data.get("responder_type"),
        "fire_station_name": data.get("fire_station_name") or "",
        "alarm_level": ALARM_LEVEL_MAP.get(
            str(data.get("alarm_level") or "").strip().upper(), data.get("alarm_level")
        ),
        "general_category": data.get("classification") or data.get("classification_of_involved"),
        "sub_category": data.get("category") or data.get("type_of_involved_general_category"),
        "fire_origin": data.get("origin") or data.get("area_of_origin"),
        "extent_of_damage": normalized_extent,
        "extent_description": data.get("extent_description") or "",
        "extent_objects_count": _safe_int(data.get("extent_objects_count"), default=None)
        if data.get("extent_objects_count") is not None
        else None,
        "stage_of_fire": data.get("stage") or data.get("stage_of_fire_upon_arrival"),
        "structures_affected": _safe_int(
            data.get("struct_aff")
            if data.get("struct_aff") is not None
            else data.get("structures_affected")
        ),
        "households_affected": _safe_int(data.get("house_aff")),
        "families_affected": _safe_int(data.get("fam_aff")),
        "individuals_affected": _safe_int(data.get("indiv_aff")),
        "vehicles_affected": _safe_int(data.get("vehic_aff")),
        "distance_from_station_km": _safe_float(
            data.get("distance_km")
            if data.get("distance_km") is not None
            else data.get("distance_from_station_km")
        ),
        "total_response_time_minutes": _safe_int(data.get("response_time")),
        "total_gas_consumed_liters": _safe_float(data.get("gas_liters")),
        "extent_total_floor_area_sqm": _safe_float(data.get("extent_total_floor_area_sqm")),
        "extent_total_land_area_hectares": _safe_float(data.get("extent_total_land_area_hectares")),
        "resources_deployed": {
            "trucks": {
                "bfp": _safe_int(
                    data.get("res_bfp_truck")
                    if data.get("res_bfp_truck") is not None
                    else data.get("res_bfp_trucks")
                ),
                "lgu": _safe_int(data.get("res_lgu_truck")),
                "volunteer": _safe_int(data.get("res_vol_truck")),
            },
            "medical": {
                "bfp": _safe_int(data.get("res_bfp_amb")),
                "non_bfp": _safe_int(data.get("res_non_amb")),
            },
            "special_assets": {
                "rescue_bfp": _safe_int(data.get("res_bfp_resc")),
                "rescue_non_bfp": _safe_int(data.get("res_non_resc")),
                "others": str(data.get("res_others") or ""),
            },
            "tools": {
                "scba": _safe_int(data.get("tool_scba")),
                "rope": str(data.get("tool_rope") or ""),
                "ladder": _safe_int(data.get("tool_ladder")),
                "hoseline": str(data.get("tool_hose") or ""),
                "hydraulic": _safe_int(data.get("tool_hydra")),
                "others": str(data.get("tool_others") or ""),
            },
            "hydrant_distance": str(data.get("hydrant_dist") or ""),
        },
        "alarm_timeline": {
            "alarm_foua": _alarm_entry("alarm_foua"),
            "alarm_1st": _alarm_entry("alarm_1st"),
            "alarm_2nd": _alarm_entry("alarm_2nd"),
            "alarm_3rd": _alarm_entry("alarm_3rd"),
            "alarm_4th": _alarm_entry("alarm_4th"),
            "alarm_5th": _alarm_entry("alarm_5th"),
            "alarm_tf_alpha": _alarm_entry("tf_alpha"),
            "alarm_tf_bravo": _alarm_entry("tf_bravo"),
            "alarm_tf_charlie": _alarm_entry("tf_charlie"),
            "alarm_tf_delta": _alarm_entry("tf_delta"),
            "alarm_general": _alarm_entry("general"),
            "alarm_fuc": _alarm_entry("fuc"),
            "alarm_fo": _alarm_entry("fo"),
            "_engines": engines,
            "_response": {
                "time_returned_to_base": _time_str(data.get("time_returned")),
                "general_description_of_involved": data.get("description") or "",
            },
        },
        "problems_encountered": data.get("problems", []),
        "recommendations": data.get("recommendations") or "",
    }

    mapped = {
        "region_id": region_id,
        "incident_nonsensitive_details": incident_nonsensitive_details,
        "incident_sensitive_details": {
            "caller_name": c_name,
            "caller_number": c_num,
            "receiver_name": data.get("receiver") or "",
            "owner_name": data.get("owner") or "",
            "establishment_name": data.get("owner") or "",
            "street_address": data.get("address") or "",
            "landmark": data.get("landmark") or "",
            "personnel_on_duty": {
                "engine_commander": data.get("pod_commander") or "",
                "shift_in_charge": data.get("pod_shift") or "",
                "nozzleman": data.get("pod_nozzleman") or "",
                "lineman": data.get("pod_lineman") or "",
                "engine_crew": data.get("pod_crew") or "",
                "driver": data.get("pod_dpo") or "",
                "pump_operator": data.get("pod_dpo") or "",
                "safety_officer": {
                    "name": data.get("pod_safety") or "",
                    "contact": data.get("pod_safety_contact") or "",
                },
                "fire_arson_investigator": {
                    "name": data.get("pod_inv") or "",
                    "contact": data.get("pod_inv_contact") or "",
                },
            },
            "other_personnel": data.get("others_list", []),
            "casualty_details": casualty_details,
            "narrative_report": data.get("narrative") or "",
            "disposition": data.get("disposition") or "",
            "disposition_prepared_by": data.get("prepared_by") or "",
            "disposition_noted_by": data.get("noted_by") or "",
            "prepared_by_officer": data.get("prepared_by") or "",
            "noted_by_officer": data.get("noted_by") or "",
            "is_icp_present": bool(data.get("icp_present")),
            "icp_location": data.get("icp_location") or "",
        },
        "responding_unit": {
            "station_name": data.get("fire_station_name") or "",
            "engine_number": data.get("engine") or "",
            "responder_type": data.get("responder_type") or "",
            "dispatch_dt": _combine_date_and_time(notif_dt, data.get("time_dispatched")),
            "arrival_dt": _combine_date_and_time(notif_dt, data.get("time_arrived")),
            "return_dt": _combine_date_and_time(notif_dt, data.get("time_returned")),
        },
        "_extra_engines": [
            {
                "station_name": data.get("fire_station_name") or "",
                "engine_number": eng["name"],
                "responder_type": data.get("responder_type") or "",
                "dispatch_dt": _combine_date_and_time(notif_dt, eng.get("time_dispatched")),
                "arrival_dt": _combine_date_and_time(notif_dt, eng.get("time_arrived")),
                "return_dt": _combine_date_and_time(notif_dt, data.get("time_returned")),
            }
            for eng in engines[1:]  # skip first engine — already in responding_unit
        ],
        "_city_text": data.get("city") or "",
        "_province_text": data.get("province") or "",
        "_region_text": data.get("region") or "",
        "_barangay_text": data.get("barangay") or "",
    }

    if not notif_dt:
        errors.append("Missing required fields: notification_dt (Check D22/D23 in XLSX)")
    if not mapped["_city_text"]:
        errors.append("Missing required fields: _city_text (City/Municipality)")

    mapped["_form_kind"] = "STRUCTURAL_AFOR"

    status = "VALID" if not errors else "INVALID"
    return AforParsedRow(row_index=0, status=status, errors=errors, data=mapped)


# ---------------------------------------------------------------------------
# Public entry points
# ---------------------------------------------------------------------------


def parse_csv_content(content: str, region_id: int) -> tuple[list[AforParsedRow], AforFormKind]:
    """Parse either the official AFOR form-style CSV or a flat tabular CSV (structural only)."""
    rows = list(csv.reader(io.StringIO(content)))
    if _looks_like_official_afor_csv(rows):
        parser = BfpXlsxParser(CsvWorksheetAdapter(rows))
        return [parse_afor_report_data(parser.parse(), region_id)], "STRUCTURAL_AFOR"

    reader = csv.DictReader(io.StringIO(content))
    results = []
    for row in reader:
        if not any(row.values()):
            continue
        results.append(parse_afor_report_data(row, region_id))
    return results, "STRUCTURAL_AFOR"


def parse_xlsx_content(content: bytes, region_id: int) -> tuple[list[AforParsedRow], AforFormKind]:
    """Parse XLSX: detect structural vs wildland, then dispatch."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), data_only=True)
    try:
        kind = detect_afor_template_kind(wb)
        if kind is None:
            raise ValueError(
                "could not determine AFOR type. Expected either the official structural AFOR "
                "(column A: 'AFTER FIRE OPERATIONS REPORT' and 'A. RESPONSE DETAILS') or the wildland "
                "template (sheet 'WILDLAND FIRE AFOR' with section A dates in column B). "
                "See public/templates/ for sample workbooks."
            )

        if kind == "STRUCTURAL_AFOR":
            ws = _pick_structural_worksheet(wb)
            parser = BfpXlsxParser(ws)
            report_data = parser.parse()
            parsed_row = parse_afor_report_data(report_data, region_id)
            return [parsed_row], kind

        ws = _pick_wildland_worksheet(wb)
        parser = WildlandXlsxParser(ws)
        report_data = parser.parse()
        parsed_row = parse_wildland_afor_report_data(report_data, region_id)
        return [parsed_row], kind
    finally:
        wb.close()


# Re-export for convenience
__all__ = [
    "AforFormKind",
    "AforParsedRow",
    "AforParseResponse",
    "WildlandRowSource",
    "BfpXlsxParser",
    "WildlandXlsxParser",
    "detect_afor_template_kind",
    "parse_afor_report_data",
    "parse_wildland_afor_report_data",
    "parse_csv_content",
    "parse_xlsx_content",
]
