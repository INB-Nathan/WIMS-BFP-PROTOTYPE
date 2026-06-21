"""
Unified regional AFOR import — structural vs wildland detection, preview, commit.

Requires DATABASE_URL / running Postgres with WIMS schema (same as test_triage_api).
"""

from __future__ import annotations

import io
import uuid
from datetime import datetime

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

import auth
from auth import get_db_with_rls
from main import app


# ---------------------------------------------------------------------------
# Helpers — minimal XLSX bytes (openpyxl)
# ---------------------------------------------------------------------------


def _build_structural_afor_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "AFOR"
    ws["A14"] = "AFTER FIRE OPERATIONS REPORT"
    ws["A18"] = "A. RESPONSE DETAILS"
    ws["B20"] = "x"
    ws["D20"] = "Test FS"
    ws["D22"] = datetime(2025, 1, 15)
    ws["D23"] = "10:00"
    ws["D26"] = "Manila"
    ws["D42"] = "First Alarm"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_wildland_afor_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "WILDLAND FIRE AFOR"
    ws["B12"] = "AFTER FIRE OPERATIONS REPORT OF WILDLAND FIRE "
    ws["B13"] = "A. DATES AND TIMES"
    ws["D15"] = datetime(2025, 3, 10, 8, 30)
    ws["E27"] = "Direct attack on head"
    ws["D23"] = "Engine 99"
    ws["G44"] = "Brush Fire"
    ws["B44"] = "12 ha"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_ambiguous_xlsx_bytes() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Notes"
    ws["A1"] = "Not an AFOR workbook"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def client():
    return TestClient(app)


@pytest.fixture(autouse=True)
def _reset_overrides():
    yield
    app.dependency_overrides.clear()


@pytest.fixture
def db_session():
    from database import _AdminSessionLocal as _SessionLocal  # noqa: SLF001

    db = _SessionLocal()
    try:
        yield db
    finally:
        db.close()


@pytest.fixture
def regional_user_id(db_session: Session):
    """REGIONAL_ENCODER with assigned NCR region (seed)."""
    keycloak_id = uuid.uuid4()
    username = f"regional_test_{keycloak_id.hex[:8]}"
    row = db_session.execute(
        text("""
            INSERT INTO wims.users (keycloak_id, username, role, assigned_region_id)
            SELECT :kid, :username, 'REGIONAL_ENCODER', region_id
            FROM wims.ref_regions WHERE region_code = 'NCR' LIMIT 1
            RETURNING user_id
        """),
        {"kid": keycloak_id, "username": username},
    ).fetchone()
    db_session.commit()
    assert row is not None
    return row[0]


@pytest.fixture
def require_wildland_schema(db_session: Session):
    """Skip DB-dependent tests when Postgres predates wildland AFOR DDL."""
    ok = db_session.execute(
        text(
            """
            SELECT EXISTS (
                SELECT 1 FROM information_schema.tables
                WHERE table_schema = 'wims' AND table_name = 'incident_wildland_afor'
            )
            """
        )
    ).scalar()
    if not ok:
        pytest.skip(
            "wims.incident_wildland_afor missing — apply src/postgres-init/01_wims_initial.sql to the DB"
        )


@pytest.fixture
def client_regional_encoder(client: TestClient, regional_user_id, db_session: Session):
    rid = db_session.execute(
        text("SELECT assigned_region_id FROM wims.users WHERE user_id = CAST(:u AS uuid)"),
        {"u": regional_user_id},
    ).scalar()

    async def mock_regional_encoder_fixed():
        return {
            "user_id": regional_user_id,
            "keycloak_id": str(uuid.uuid4()),
            "role": "REGIONAL_ENCODER",
            "assigned_region_id": rid,
        }

    def mock_rls_db():
        try:
            yield db_session
        finally:
            pass

    app.dependency_overrides[auth.get_current_wims_user] = mock_regional_encoder_fixed
    app.dependency_overrides[auth.get_regional_encoder] = mock_regional_encoder_fixed
    app.dependency_overrides[get_db_with_rls] = mock_rls_db
    try:
        yield client
    finally:
        app.dependency_overrides.pop(auth.get_current_wims_user, None)
        app.dependency_overrides.pop(auth.get_regional_encoder, None)
        app.dependency_overrides.pop(get_db_with_rls, None)


# ---------------------------------------------------------------------------
# Helpers — WGS84 / PostGIS
# ---------------------------------------------------------------------------

# Distinct from legacy placeholder POINT(121.0 14.5) used before real coords.
SAMPLE_LAT = 14.5547
SAMPLE_LON = 121.0244


def _fetch_incident_wgs84(db_session: Session, incident_id: int) -> tuple[float, float]:
    """Return (longitude, latitude) from fire_incidents.location (SRID 4326)."""
    row = db_session.execute(
        text(
            """
            SELECT ST_X(location::geometry), ST_Y(location::geometry)
            FROM wims.fire_incidents WHERE incident_id = :id
            """
        ),
        {"id": incident_id},
    ).fetchone()
    assert row is not None
    return float(row[0]), float(row[1])


def _commit_coords_body() -> dict:
    return {"latitude": SAMPLE_LAT, "longitude": SAMPLE_LON}


def _assert_wgs84_error(res, status: int = 400):
    assert res.status_code == status
    detail = res.json().get("detail")
    if isinstance(detail, dict):
        assert detail.get("code") == "AFOR_WGS84_INVALID"
        assert "message" in detail
    else:
        assert isinstance(detail, str)
        assert "WGS84" in detail or "latitude" in detail.lower()


# ---------------------------------------------------------------------------
# Tests
# ---------------------------------------------------------------------------


def test_regional_import_preview_structural_form_kind(
    client_regional_encoder: TestClient,
):
    response = client_regional_encoder.post(
        "/api/regional/afor/import",
        files={
            "file": (
                "struct.xlsx",
                _build_structural_afor_xlsx_bytes(),
                "application/octet-stream",
            )
        },
    )
    assert response.status_code == 200
    data = response.json()
    assert data.get("form_kind") == "STRUCTURAL_AFOR"
    assert data.get("requires_location") is True


def test_regional_import_rejects_deprecated_wildland_afor(
    client_regional_encoder: TestClient,
):
    response = client_regional_encoder.post(
        "/api/regional/afor/import",
        files={
            "file": (
                "wild.xlsx",
                _build_wildland_afor_xlsx_bytes(),
                "application/octet-stream",
            )
        },
    )
    assert response.status_code == 400
    detail = response.json().get("detail", "")
    assert "Wildland-specific AFOR import is deprecated" in detail


def test_commit_rejects_deprecated_wildland_afor(
    require_wildland_schema,
    client_regional_encoder: TestClient,
):
    row = {
        "_form_kind": "WILDLAND_AFOR",
        "_city_text": "",
        "region_id": 1,
        "wildland": {
            "primary_action_taken": "Direct attack on head",
            "engine_dispatched": "Engine 42",
            "wildland_fire_type": "brush fire",
        },
    }
    res = client_regional_encoder.post(
        "/api/regional/afor/commit",
        json={"form_kind": "WILDLAND_AFOR", "rows": [row], **_commit_coords_body()},
    )
    assert res.status_code == 400
    detail = res.json().get("detail", "")
    assert "Wildland-specific AFOR commit is deprecated" in detail


def test_commit_structural_persists_wgs84_coordinates(
    client_regional_encoder: TestClient,
    db_session: Session,
):
    prev = client_regional_encoder.post(
        "/api/regional/afor/import",
        files={
            "file": (
                "struct.xlsx",
                _build_structural_afor_xlsx_bytes(),
                "application/octet-stream",
            )
        },
    )
    assert prev.status_code == 200
    rows = [r["data"] for r in prev.json()["rows"] if r["status"] == "VALID"]
    assert rows
    commit = client_regional_encoder.post(
        "/api/regional/afor/commit",
        json={"form_kind": "STRUCTURAL_AFOR", "rows": rows, **_commit_coords_body()},
    )
    assert commit.status_code == 200, commit.text
    # If the test DB has seed data that matches our minimal XLSX, the backend
    # returns DUPLICATE_CHECK_REQUIRED. Re-commit with a "force" resolution so
    # the test can verify WGS84 persistence regardless of pre-existing duplicates.
    if commit.status_code == 200 and commit.json().get("status") == "DUPLICATE_CHECK_REQUIRED":
        commit = client_regional_encoder.post(
            "/api/regional/afor/commit",
            json={
                "form_kind": "STRUCTURAL_AFOR",
                "rows": rows,
                **_commit_coords_body(),
                "resolutions": [{"row_index": 0, "action": "force"}],
            },
        )
        assert commit.status_code == 200, commit.text
    iid = commit.json()["incident_ids"][0]
    lon, lat = _fetch_incident_wgs84(db_session, iid)
    assert abs(lon - SAMPLE_LON) < 1e-5
    assert abs(lat - SAMPLE_LAT) < 1e-5


def test_commit_rejects_form_kind_mismatch(
    require_wildland_schema,
    client_regional_encoder: TestClient,
):
    prev = client_regional_encoder.post(
        "/api/regional/afor/import",
        files={
            "file": (
                "wild.xlsx",
                _build_wildland_afor_xlsx_bytes(),
                "application/octet-stream",
            )
        },
    )
    assert prev.status_code == 200
    rows = [r["data"] for r in prev.json()["rows"] if r["status"] == "VALID"]

    res = client_regional_encoder.post(
        "/api/regional/afor/commit",
        json={"form_kind": "STRUCTURAL_AFOR", "rows": rows, **_commit_coords_body()},
    )
    assert res.status_code == 400
