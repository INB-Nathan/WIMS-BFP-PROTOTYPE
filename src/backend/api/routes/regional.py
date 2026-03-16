"""Regional Office API — AFOR Import, Region-Scoped Incidents, Stats.

All endpoints protected by get_regional_encoder (REGIONAL_ENCODER role + assigned_region_id).
Data isolation: every query filters by the user's assigned_region_id.
"""

from __future__ import annotations

import csv
import io
import logging
import uuid
from datetime import datetime
from typing import Annotated, Any, Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile
from pydantic import BaseModel
from sqlalchemy import text
from sqlalchemy.orm import Session

from auth import get_regional_encoder
from database import get_db

logger = logging.getLogger("wims.regional")

router = APIRouter(prefix="/api/regional", tags=["regional"])


# ---------------------------------------------------------------------------
# Pydantic Schemas
# ---------------------------------------------------------------------------

class AforParsedRow(BaseModel):
    row_index: int
    status: str  # VALID | INVALID
    errors: list[str]
    data: dict[str, Any]


class AforParseResponse(BaseModel):
    total_rows: int
    valid_rows: int
    invalid_rows: int
    rows: list[AforParsedRow]


class AforCommitRequest(BaseModel):
    rows: list[dict[str, Any]]


class AforCommitResponse(BaseModel):
    status: str
    batch_id: int
    incident_ids: list[int]
    total_committed: int


class RegionalStatsResponse(BaseModel):
    total_incidents: int
    by_category: list[dict[str, Any]]
    by_alarm_level: list[dict[str, Any]]
    by_status: list[dict[str, Any]]


# ---------------------------------------------------------------------------
# AFOR Parsing Utilities (Official BFP XLSX Refactor)
# ---------------------------------------------------------------------------

ALARM_LEVEL_MAP = {
    "1ST": "First Alarm",
    "1ST ALARM": "First Alarm",
    "FIRST": "First Alarm",
    "FIRST ALARM": "First Alarm",
    "2ND": "Second Alarm",
    "2ND ALARM": "Second Alarm",
    "SECOND": "Second Alarm",
    "SECOND ALARM": "Second Alarm",
    "3RD": "Third Alarm",
    "3RD ALARM": "Third Alarm",
    "THIRD": "Third Alarm",
    "THIRD ALARM": "Third Alarm",
    "4TH": "Fourth Alarm",
    "4TH ALARM": "Fourth Alarm",
    "FOURTH": "Fourth Alarm",
    "FOURTH ALARM": "Fourth Alarm",
    "5TH": "Fifth Alarm",
    "5TH ALARM": "Fifth Alarm",
    "FIFTH": "Fifth Alarm",
    "FIFTH ALARM": "Fifth Alarm",
    "TF ALPHA": "Task Force Alpha",
    "TASK FORCE ALPHA": "Task Force Alpha",
    "TF BRAVO": "Task Force Bravo",
    "TASK FORCE BRAVO": "Task Force Bravo",
    "TF CHARLIE": "Task Force Charlie",
    "TASK FORCE CHARLIE": "Task Force Charlie",
    "TF DELTA": "Task Force Delta",
    "TASK FORCE DELTA": "Task Force Delta",
    "GENERAL": "General Alarm",
    "GENERAL ALARM": "General Alarm",
}


def _safe_int(val: Any, default: int = 0) -> int:
    if val is None or val == "" or val == "N/A":
        return default
    try:
        if isinstance(val, (int, float)):
            return int(val)
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return default


def _safe_float(val: Any, default: float = 0.0) -> float:
    if val is None or val == "" or val == "N/A":
        return default
    try:
        return float(str(val).strip())
    except (ValueError, TypeError):
        return default


def _safe_dt(val: Any) -> str | None:
    """Safe datetime string conversion."""
    if isinstance(val, datetime):
        return val.isoformat()
    if not val:
        return None
    
    dt_str = str(val).strip()
    for fmt in ["%Y-%m-%d %H:%M", "%m-%d-%Y %H:%M", "%H:%M", "%Y-%m-%d", "%m-%d-%Y", "%m/%d/%Y"]:
        try:
            return datetime.strptime(dt_str, fmt).isoformat()
        except ValueError:
            continue
    return None


class BfpXlsxParser:
    """Parser for the official BFP manual entry form (AFOR)."""
    
    def __init__(self, ws):
        self.ws = ws

    def get(self, coord: str) -> Any:
        val = self.ws[coord].value
        if val is None: return None
        if isinstance(val, str): return val.strip()
        return val

    def _is_marked(self, coord: str) -> bool:
        val = str(self.get(coord)).strip().lower() if self.get(coord) else ""
        return val in ['x', '1', 'true', 'v', '✓', '✔', '/']

    def parse(self) -> dict[str, Any]:
        """Extract sections A through L into a comprehensive data dictionary."""
        
        # Section A: Response Details
        responder_type = "First Responder" if self._is_marked("B20") else ("Augmenting Team" if self._is_marked("B21") else "First Responder")
        
        # Section B: Classification
        classification = "Structural"
        cat_val = self.get("D48")
        if self._is_marked("B49"):
            classification = "Non-Structural"
            cat_val = self.get("D49")
        elif self._is_marked("B50"):
            classification = "Transportation"
            cat_val = self.get("D50")

        # Extent of Damage
        extent = "None / Minor"
        if self._is_marked("C57"): extent = "Confined to Object"
        elif self._is_marked("C58"): extent = "Confined to Room"
        elif self._is_marked("B59"): extent = "Confined to Structure"
        elif self._is_marked("C60"): extent = "Total Loss"
        elif self._is_marked("C61"): extent = "Extended Beyond Structure"

        # Section J: Problems
        problems = []
        prob_map = {
            "C195": "Inaccurate address", "C196": "Geographically challenged",
            "C197": "Road conditions", "C198": "Road under construction",
            "B199": "Traffic congestion", "C200": "Road accidents",
            "C201": "Vehicles failure to yield", "C202": "Natural Disasters",
            "C203": "Civil Disturbance", "B210": "Intense heat and smoke"
        }
        for c, flavor in prob_map.items():
            if self._is_marked(c): problems.append(flavor)

        # Section I: Narrative joining (Rows 160 to 190)
        narrative_lines = []
        for r in range(160, 191):
            line = self.get(f"B{r}")
            if line: narrative_lines.append(str(line))
        
        # Section G: Other Personnel (Rows 124 to 132)
        others = []
        for r in range(124, 133):
            name = self.get(f"B{r}")
            rem = self.get(f"E{r}")
            if name and "N/A" not in str(name).upper():
                others.append({"name": name, "designation": rem or ""})

        return {
            "responder_type": responder_type,
            "fire_station_name": self.get("D20") if responder_type == "First Responder" else self.get("D21"),
            "notification_date": self.get("D22"),
            "notification_time": self.get("D23"),
            "region": self.get("D24"),
            "province": self.get("D25"),
            "city": self.get("D26"),
            "address": self.get("D27"),
            "landmark": self.get("D28"),
            "caller_info": self.get("D29"),
            "receiver": self.get("D30"),
            "engine": self.get("D31"),
            "time_dispatched": self.get("D34"),
            "time_arrived": self.get("D37"),
            "response_time": self.get("D40"),
            "distance_km": self.get("D41"),
            "alarm_level": self.get("D42"),
            "time_returned": self.get("D43"),
            "gas_liters": self.get("D44"),

            "classification": classification,
            "category": cat_val,
            "owner": self.get("D51"),
            "description": self.get("D52"),
            "origin": self.get("D53"),
            "stage": self.get("D54"),
            "extent": extent,
            
            "extent_total_floor_area_sqm": self.get("D56") or self.get("D57") or self.get("D58") or self.get("D59") or self.get("D60"),
            "extent_total_land_area_hectares": self.get("D59") or self.get("D60"),
            
            "struct_aff": self.get("D62"),
            "house_aff": self.get("D63"),
            "fam_aff": self.get("D64"),
            "indiv_aff": self.get("D65"),
            "vehic_aff": self.get("D66"),

            "res_bfp_truck": self.get("D70"),
            "res_lgu_truck": self.get("D71"),
            "res_vol_truck": self.get("D72"),
            "res_bfp_amb": self.get("D73"),
            "res_non_amb": self.get("D74"),
            "res_bfp_resc": self.get("D75"),
            "res_non_resc": self.get("D76"),
            "res_others": self.get("D77"),
            "tool_scba": self.get("D79"),
            "tool_rope": self.get("D80"),
            "tool_ladder": self.get("D81"),
            "tool_hose": self.get("D82"),
            "tool_hydra": self.get("D83"),
            "tool_others": self.get("D84"),
            "hydrant_dist": self.get("D85"),

            "timeline": {
                "alarm_1st": {"time": self.get("D89"), "date": self.get("E89")},
                "alarm_2nd": {"time": self.get("D90"), "date": self.get("E90")},
                "alarm_3rd": {"time": self.get("D91"), "date": self.get("E91")},
                "alarm_4th": {"time": self.get("D92"), "date": self.get("E92")},
                "alarm_5th": {"time": self.get("D93"), "date": self.get("E93")},
                "tf_alpha": {"time": self.get("D94"), "date": self.get("E94")},
                "tf_bravo": {"time": self.get("D95"), "date": self.get("E95")},
                "tf_charlie": {"time": self.get("D96"), "date": self.get("E96")},
                "tf_delta": {"time": self.get("D97"), "date": self.get("E97")},
                "general": {"time": self.get("D98"), "date": self.get("E98")},
                "fuc": {"time": self.get("D99"), "date": self.get("E99")},
                "fo": {"time": self.get("D100"), "date": self.get("E100")}
            },
            
            "inj_civ_m": self.get("D106"), "inj_civ_f": self.get("E106"),
            "inj_bfp_m": self.get("D107"), "inj_bfp_f": self.get("E107"),
            "inj_aux_m": self.get("D108"), "inj_aux_f": self.get("E108"),
            "fat_civ_m": self.get("D109"), "fat_civ_f": self.get("E109"),
            "fat_bfp_m": self.get("D110"), "fat_bfp_f": self.get("E110"),
            "fat_aux_m": self.get("D111"), "fat_aux_f": self.get("E111"),

            "pod_commander": self.get("D114"),
            "pod_shift": self.get("D115"),
            "pod_nozzleman": self.get("D116"),
            "pod_lineman": self.get("D117"),
            "pod_crew": self.get("D118"),
            "pod_dpo": self.get("D119"),
            "pod_safety": self.get("D120"),
            
            "others_list": others,
            "narrative": "\n".join(narrative_lines),
            "problems": problems,
            "recommendations": self.get("B222"),
            "disposition": self.get("B229"),
            "prepared_by": self.get("C238"),
            "noted_by": self.get("F238")
        }


def parse_afor_report_data(data: dict, region_id: int) -> AforParsedRow:
    """Map the extracted AFOR dictionary into the strict database schema."""
    errors: list[str] = []
    
    def _dt(d, t=None):
        if not d: return None
        try:
            from datetime import date, time
            # Handle date
            if isinstance(d, date):
                d_str = d.strftime("%Y-%m-%d")
            else:
                d_str = str(d).split(" ")[0]
            
            # Handle time
            if t:
                if isinstance(t, time):
                    t_str = t.strftime("%H:%M:%S")
                else:
                    # Excel might give a float represent of time or a string
                    t_str = str(t)
                return f"{d_str}T{t_str}"
            return f"{d_str}T00:00:00"
        except:
            return None

    notif_dt = _dt(data.get("notification_date"), data.get("notification_time"))
    
    # Split caller_info: "Name / 0917-..."
    ci = str(data.get("caller_info") or "")
    c_name = ci.split("/")[0].strip() if "/" in ci else ci
    c_num = ci.split("/")[1].strip() if "/" in ci else ""

    mapped = {
        "region_id": region_id,
        "incident_nonsensitive_details": {
            "notification_dt": notif_dt, # Return None if not parsed, let frontend/DB handle fallback
            "responder_type": data.get("responder_type"),
            "fire_station_name": data.get("fire_station_name") or "",
            "region": data.get("region"),
            "province_district": data.get("province"),
            "city_municipality": data.get("city"),
            "incident_address": data.get("address"),
            "nearest_landmark": data.get("landmark"),
            "receiver_name": data.get("receiver"),
            "engine_dispatched": data.get("engine"),
            "time_engine_dispatched": str(data.get("time_dispatched")) if data.get("time_dispatched") else "",
            "time_arrived_at_scene": str(data.get("time_arrived")) if data.get("time_arrived") else "",
            "total_response_time_minutes": _safe_int(data.get("response_time")),
            "distance_to_fire_scene_km": _safe_float(data.get("distance_km")),
            "alarm_level": ALARM_LEVEL_MAP.get(str(data.get("alarm_level") or "").strip().upper(), data.get("alarm_level")),
            "time_returned_to_base": str(data.get("time_returned")) if data.get("time_returned") else "",
            "total_gas_consumed_liters": _safe_float(data.get("gas_liters")),
            
            "classification_of_involved": data.get("classification"),
            "type_of_involved_general_category": data.get("category"),
            "owner_name": data.get("owner"),
            "general_description_of_involved": data.get("description"),
            "area_of_origin": data.get("origin"),
            "stage_of_fire_upon_arrival": data.get("stage"),
            "extent_of_damage": data.get("extent"),
            "extent_total_floor_area_sqm": _safe_float(data.get("extent_total_floor_area_sqm")),
            "extent_total_land_area_hectares": _safe_float(data.get("extent_total_land_area_hectares")),
            
            "structures_affected": _safe_int(data.get("struct_aff")),
            "households_affected": _safe_int(data.get("house_aff")),
            "families_affected": _safe_int(data.get("fam_aff")),
            "individuals_affected": _safe_int(data.get("indiv_aff")),
            "vehicles_affected": _safe_int(data.get("vehic_aff")),

            "resources_deployed": {
                "trucks": {
                    "bfp": _safe_int(data.get("res_bfp_truck")),
                    "lgu": _safe_int(data.get("res_lgu_truck")),
                    "volunteer": _safe_int(data.get("res_vol_truck"))
                },
                "medical": {
                    "bfp": _safe_int(data.get("res_bfp_amb")),
                    "non_bfp": _safe_int(data.get("res_non_amb"))
                },
                "special_assets": {
                    "rescue_bfp": _safe_int(data.get("res_bfp_resc")),
                    "rescue_non_bfp": _safe_int(data.get("res_non_resc")),
                    "others": str(data.get("res_others") or "")
                },
                "tools": {
                    "scba": _safe_int(data.get("tool_scba")),
                    "rope": str(data.get("tool_rope") or ""),
                    "ladder": _safe_int(data.get("tool_ladder")),
                    "hoseline": str(data.get("tool_hose") or ""),
                    "hydraulic": _safe_int(data.get("tool_hydra")),
                    "others": str(data.get("tool_others") or "")
                },
                "hydrant_distance": str(data.get("hydrant_dist") or "")
            },
            
            "alarm_timeline": {
                "alarm_1st": _dt(data["timeline"]["alarm_1st"]["date"], data["timeline"]["alarm_1st"]["time"]),
                "alarm_2nd": _dt(data["timeline"]["alarm_2nd"]["date"], data["timeline"]["alarm_2nd"]["time"]),
                "alarm_3rd": _dt(data["timeline"]["alarm_3rd"]["date"], data["timeline"]["alarm_3rd"]["time"]),
                "alarm_4th": _dt(data["timeline"]["alarm_4th"]["date"], data["timeline"]["alarm_4th"]["time"]),
                "alarm_5th": _dt(data["timeline"]["alarm_5th"]["date"], data["timeline"]["alarm_5th"]["time"]),
                "alarm_tf_alpha": _dt(data["timeline"]["tf_alpha"]["date"], data["timeline"]["tf_alpha"]["time"]),
                "alarm_tf_bravo": _dt(data["timeline"]["tf_bravo"]["date"], data["timeline"]["tf_bravo"]["time"]),
                "alarm_tf_charlie": _dt(data["timeline"]["tf_charlie"]["date"], data["timeline"]["tf_charlie"]["time"]),
                "alarm_tf_delta": _dt(data["timeline"]["tf_delta"]["date"], data["timeline"]["tf_delta"]["time"]),
                "alarm_general": _dt(data["timeline"]["general"]["date"], data["timeline"]["general"]["time"]),
                "alarm_fuc": _dt(data["timeline"]["fuc"]["date"], data["timeline"]["fuc"]["time"]),
                "alarm_fo": _dt(data["timeline"]["fo"]["date"], data["timeline"]["fo"]["time"])
            },

            "casualty_details": {
                "injured": {
                    "civilian_m": _safe_int(data.get("inj_civ_m")),
                    "civilian_f": _safe_int(data.get("inj_civ_f")),
                    "bfp_m": _safe_int(data.get("inj_bfp_m")),
                    "bfp_f": _safe_int(data.get("inj_bfp_f")),
                    "aux_m": _safe_int(data.get("inj_aux_m")),
                    "aux_f": _safe_int(data.get("inj_aux_f")),
                },
                "fatalities": {
                    "civilian_m": _safe_int(data.get("fat_civ_m")),
                    "civilian_f": _safe_int(data.get("fat_civ_f")),
                    "bfp_m": _safe_int(data.get("fat_bfp_m")),
                    "bfp_f": _safe_int(data.get("fat_bfp_f")),
                    "aux_m": _safe_int(data.get("fat_aux_m")),
                    "aux_f": _safe_int(data.get("fat_aux_f")),
                }
            },
            "problems_encountered": data.get("problems", []),
            "recommendations": data.get("recommendations") or "",
            "other_personnel": data.get("others_list", [])
        },
        "incident_sensitive_details": {
            "caller_name": c_name,
            "caller_number": c_num,
            "receiver_name": data.get("receiver") or "",
            "owner_name": data.get("owner") or "",
            "establishment_name": data.get("owner") or "",
            "street_address": data.get("address") or "",
            "landmark": data.get("landmark") or "",
            
            "personnel_on_duty": {
                "engine_commander": data.get("pod_commander") or "",
                "shift_in_charge": data.get("pod_shift") or "",
                "nozzleman": data.get("pod_nozzleman") or "",
                "lineman": data.get("pod_lineman") or "",
                "engine_crew": data.get("pod_crew") or "",
                "driver": data.get("pod_dpo") or "",
                "pump_operator": data.get("pod_dpo") or "",
                "safety_officer": {"name": data.get("pod_safety") or "", "contact": ""}
            }
        },
        "narrative_report": data.get("narrative"),
        "recommendations": data.get("recommendations"),
        "disposition": data.get("disposition"),
        "prepared_by": data.get("prepared_by") or "",
        "noted_by": data.get("noted_by") or "",
        "_city_text": data.get("city") or "",
        "_province_text": data.get("province") or "",
    }
    
    if not notif_dt:
        errors.append("Missing required fields: notification_dt (Check D22/D23 in XLSX)")
    if not mapped["_city_text"]:
        errors.append("Missing required fields: _city_text (City/Municipality)")
    
    status = "VALID" if not errors else "INVALID"
    return AforParsedRow(row_index=0, status=status, errors=errors, data=mapped)


def parse_csv_content(content: str, region_id: int) -> list[AforParsedRow]:
    """Parse legacy CSV or flat tabular CSV data."""
    # (Existing flat parser logic as fallback or for bulk tabular imports)
    reader = csv.DictReader(io.StringIO(content))
    results = []
    # For CSV we assume it's the simplified flat format or a single-row export
    for idx, row in enumerate(reader):
        if not any(row.values()): continue
        # For CSV we can't easily map the BFP form sections, so we use canonical keys
        # This implementation is kept as a fallback if the user uploads a flat CSV
        results.append(parse_afor_report_data(row, region_id))
    return results


def parse_xlsx_content(content: bytes, region_id: int) -> list[AforParsedRow]:
    """Parse XLSX using BfpXlsxParser for official BFP template."""
    from openpyxl import load_workbook
    
    wb = load_workbook(io.BytesIO(content), data_only=True)
    # Target 'AFOR' sheet
    ws = None
    for name in wb.sheetnames:
        if "AFOR" in name.upper():
            ws = wb[name]
            break
    
    if not ws:
        # Fallback to active if no AFOR sheet found
        ws = wb.active
        
    parser = BfpXlsxParser(ws)
    report_data = parser.parse()
    parsed_row = parse_afor_report_data(report_data, region_id)
    
    wb.close()
    return [parsed_row]


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.post("/afor/import", response_model=AforParseResponse)
async def import_afor_file(
    file: UploadFile = File(...),
    user: dict = Depends(get_regional_encoder),
    db: Session = Depends(get_db),
):
    """
    Upload and parse an AFOR file (.xlsx or .csv).
    Returns parsed rows with validation status for preview before commit.
    """
    region_id = user["assigned_region_id"]

    if not file.filename:
        raise HTTPException(status_code=400, detail="No file provided")

    ext = file.filename.rsplit(".", 1)[-1].lower() if "." in file.filename else ""
    if ext not in ("xlsx", "xls", "csv"):
        raise HTTPException(status_code=400, detail="Only .xlsx, .xls, and .csv files are supported")

    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="File is empty")

    try:
        if ext == "csv":
            decoded = content.decode("utf-8-sig")  # Handle BOM
            rows = parse_csv_content(decoded, region_id)
        else:
            rows = parse_xlsx_content(content, region_id)
    except Exception as e:
        logger.error(f"Failed to parse AFOR file: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse file: {str(e)}")

    if len(rows) == 0:
        raise HTTPException(status_code=400, detail="No data rows found in file")

    valid_count = sum(1 for r in rows if r.status == "VALID")

    return AforParseResponse(
        total_rows=len(rows),
        valid_rows=valid_count,
        invalid_rows=len(rows) - valid_count,
        rows=rows,
    )


@router.post("/afor/commit", response_model=AforCommitResponse)
def commit_afor_import(
    body: AforCommitRequest,
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db)],
):
    """
    Commit validated AFOR rows to the database.
    Creates a data_import_batch and inserts fire_incidents with details.
    """
    region_id = user["assigned_region_id"]
    user_id = user["user_id"]

    if not body.rows:
        raise HTTPException(status_code=400, detail="No rows to commit")

    # Create import batch
    batch_row = db.execute(
        text("""
            INSERT INTO wims.data_import_batches (region_id, uploaded_by, record_count)
            VALUES (:region_id, CAST(:uid AS uuid), :count)
            RETURNING batch_id
        """),
        {"region_id": region_id, "uid": user_id, "count": len(body.rows)},
    ).fetchone()

    if not batch_row:
        raise HTTPException(status_code=500, detail="Failed to create import batch")

    batch_id = batch_row[0]
    incident_ids: list[int] = []

    for row_data in body.rows:
        ns = row_data.get("incident_nonsensitive_details", {})
        sens = row_data.get("incident_sensitive_details", {})

        # Insert fire_incident
        inc_row = db.execute(
            text("""
                INSERT INTO wims.fire_incidents
                    (import_batch_id, encoder_id, region_id, location, verification_status)
                VALUES
                    (:batch_id, CAST(:uid AS uuid), :region_id,
                     ST_GeogFromText('SRID=4326;POINT(121.0 14.5)'),
                     'DRAFT')
                RETURNING incident_id
            """),
            {"batch_id": batch_id, "uid": user_id, "region_id": region_id},
        ).fetchone()

        if not inc_row:
            continue

        incident_id = inc_row[0]
        incident_ids.append(incident_id)

        # Resolve Geography IDs
        city_text = row_data.get("_city_text", "")
        geo_ids = db.execute(
            text("""
                SELECT c.city_id, c.province_id 
                FROM wims.ref_cities c
                WHERE LOWER(c.city_name) = LOWER(:city)
                LIMIT 1
            """),
            {"city": city_text}
        ).fetchone()
        
        city_id = geo_ids[0] if geo_ids else None
        province_id = geo_ids[1] if geo_ids else None

        # Insert nonsensitive details
        import json
        db.execute(
            text("""
                INSERT INTO wims.incident_nonsensitive_details (
                    incident_id, notification_dt, responder_type, fire_station_name,
                    alarm_level, general_category, sub_category,
                    city_id, province_id, district_id,
                    fire_origin, extent_of_damage, stage_of_fire,
                    structures_affected, households_affected, families_affected,
                    individuals_affected, vehicles_affected,
                    total_response_time_minutes, total_gas_consumed_liters,
                    extent_total_floor_area_sqm, extent_total_land_area_hectares,
                    civilian_injured, civilian_deaths, firefighter_injured, firefighter_deaths,
                    resources_deployed, alarm_timeline, problems_encountered,
                    recommendations
                ) VALUES (
                    :incident_id, CAST(:notification_dt AS timestamptz),
                    :responder_type, :fire_station_name,
                    :alarm_level, :general_category, :sub_category,
                    :city_id, :province_id, 1,
                    :fire_origin, :extent_of_damage, :stage_of_fire,
                    :structures_affected, :households_affected, :families_affected,
                    :individuals_affected, :vehicles_affected,
                    :total_response_time_minutes, :total_gas_consumed_liters,
                    :floor_area, :land_area,
                    :civ_inj, :civ_fat, :ff_inj, :ff_fat,
                    CAST(:resources_deployed AS jsonb), CAST(:alarm_timeline AS jsonb),
                    CAST(:problems_encountered AS jsonb),
                    :recommendations
                )
            """),
            {
                "incident_id": incident_id,
                "notification_dt": ns.get("notification_dt"),
                "responder_type": ns.get("responder_type", ""),
                "fire_station_name": ns.get("fire_station_name", ""),
                "alarm_level": ns.get("alarm_level", ""),
                "general_category": ns.get("general_category", ""),
                "sub_category": ns.get("incident_type", ""),
                "city_id": city_id,
                "province_id": province_id,
                "fire_origin": ns.get("fire_origin", ""),
                "extent_of_damage": ns.get("extent_of_damage", ""),
                "stage_of_fire": ns.get("stage_of_fire", ""),
                "structures_affected": ns.get("structures_affected", 0),
                "households_affected": ns.get("households_affected", 0),
                "families_affected": ns.get("families_affected", 0),
                "individuals_affected": ns.get("individuals_affected", 0),
                "vehicles_affected": ns.get("vehicles_affected", 0),
                "total_response_time_minutes": ns.get("total_response_time_minutes", 0),
                "total_gas_consumed_liters": ns.get("total_gas_consumed_liters", 0),
                "floor_area": ns.get("extent_total_floor_area_sqm", 0),
                "land_area": ns.get("extent_total_land_area_hectares", 0),
                "civ_inj": sens.get("casualty_details", {}).get("injured", {}).get("civilian", {}).get("m", 0) + sens.get("casualty_details", {}).get("injured", {}).get("civilian", {}).get("f", 0),
                "civ_fat": sens.get("casualty_details", {}).get("fatal", {}).get("civilian", {}).get("m", 0) + sens.get("casualty_details", {}).get("fatal", {}).get("civilian", {}).get("f", 0),
                "ff_inj": sens.get("casualty_details", {}).get("injured", {}).get("firefighter", {}).get("m", 0) + sens.get("casualty_details", {}).get("injured", {}).get("firefighter", {}).get("f", 0),
                "ff_fat": sens.get("casualty_details", {}).get("fatal", {}).get("firefighter", {}).get("m", 0) + sens.get("casualty_details", {}).get("fatal", {}).get("firefighter", {}).get("f", 0),
                "resources_deployed": json.dumps(ns.get("resources_deployed", {})),
                "alarm_timeline": json.dumps(ns.get("alarm_timeline", {})),
                "problems_encountered": json.dumps(ns.get("problems_encountered", [])),
                "recommendations": ns.get("recommendations", ""),
            },
        )

        # Insert sensitive details
        db.execute(
            text("""
                INSERT INTO wims.incident_sensitive_details (
                    incident_id, street_address, landmark,
                    caller_name, caller_number, receiver_name,
                    owner_name, establishment_name,
                    narrative_report, disposition,
                    disposition_prepared_by, disposition_noted_by,
                    prepared_by_officer, noted_by_officer,
                    personnel_on_duty, other_personnel, casualty_details,
                    is_icp_present, icp_location
                ) VALUES (
                    :incident_id, :street_address, :landmark,
                    :caller_name, :caller_number, :receiver_name,
                    :owner_name, :establishment_name,
                    :narrative_report, :disposition,
                    :disposition_prepared_by, :disposition_noted_by,
                    :disposition_prepared_by, :disposition_noted_by,
                    CAST(:personnel_on_duty AS jsonb),
                    CAST(:other_personnel AS jsonb),
                    CAST(:casualty_details AS jsonb),
                    :is_icp_present, :icp_location
                )
            """),
            {
                "incident_id": incident_id,
                "street_address": sens.get("street_address", ""),
                "landmark": sens.get("landmark", ""),
                "caller_name": sens.get("caller_name", ""),
                "caller_number": sens.get("caller_number", ""),
                "receiver_name": sens.get("receiver_name", ""),
                "owner_name": sens.get("owner_name", ""),
                "establishment_name": sens.get("establishment_name", ""),
                "narrative_report": sens.get("narrative_report", ""),
                "disposition": sens.get("disposition", ""),
                "disposition_prepared_by": sens.get("disposition_prepared_by", ""),
                "disposition_noted_by": sens.get("disposition_noted_by", ""),
                "personnel_on_duty": json.dumps(sens.get("personnel_on_duty", {})),
                "other_personnel": json.dumps(sens.get("other_personnel", [])),
                "casualty_details": json.dumps(sens.get("casualty_details", {})),
                "is_icp_present": sens.get("is_icp_present", False),
                "icp_location": sens.get("icp_location", ""),
            },
        )

    db.commit()

    return AforCommitResponse(
        status="ok",
        batch_id=batch_id,
        incident_ids=incident_ids,
        total_committed=len(incident_ids),
    )


@router.get("/incidents")
def get_regional_incidents(
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db)],
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    category: Optional[str] = None,
    status: Optional[str] = None,
):
    """
    Fetch fire incidents scoped to the user's assigned region.
    Joins nonsensitive details for summary view.
    """
    region_id = user["assigned_region_id"]

    where_clauses = ["fi.region_id = :region_id", "fi.is_archived = FALSE"]
    params: dict[str, Any] = {"region_id": region_id, "limit": limit, "offset": offset}

    if category:
        where_clauses.append("nd.general_category = :category")
        params["category"] = category
    if status:
        where_clauses.append("fi.verification_status = :status")
        params["status"] = status

    where_sql = " AND ".join(where_clauses)

    rows = db.execute(
        text(f"""
            SELECT fi.incident_id, fi.verification_status, fi.created_at,
                   nd.notification_dt, nd.general_category, nd.alarm_level,
                   nd.fire_station_name, nd.structures_affected,
                   nd.households_affected, nd.individuals_affected,
                   nd.responder_type, nd.fire_origin, nd.extent_of_damage,
                   sd.owner_name, sd.establishment_name, sd.caller_name
            FROM wims.fire_incidents fi
            LEFT JOIN wims.incident_nonsensitive_details nd ON nd.incident_id = fi.incident_id
            LEFT JOIN wims.incident_sensitive_details sd ON sd.incident_id = fi.incident_id
            WHERE {where_sql}
            ORDER BY fi.created_at DESC
            LIMIT :limit OFFSET :offset
        """),
        params,
    ).fetchall()

    total = db.execute(
        text(f"""
            SELECT COUNT(*) FROM wims.fire_incidents fi
            LEFT JOIN wims.incident_nonsensitive_details nd ON nd.incident_id = fi.incident_id
            WHERE {where_sql}
        """),
        {k: v for k, v in params.items() if k not in ("limit", "offset")},
    ).scalar() or 0

    return {
        "items": [
            {
                "incident_id": r[0],
                "verification_status": r[1],
                "created_at": r[2].isoformat() if r[2] else None,
                "notification_dt": r[3].isoformat() if r[3] else None,
                "general_category": r[4],
                "alarm_level": r[5],
                "fire_station_name": r[6],
                "structures_affected": r[7],
                "households_affected": r[8],
                "individuals_affected": r[9],
                "responder_type": r[10],
                "fire_origin": r[11],
                "extent_of_damage": r[12],
                "owner_name": r[13],
                "establishment_name": r[14],
                "caller_name": r[15],
            }
            for r in rows
        ],
        "total": total,
        "limit": limit,
        "offset": offset,
    }


@router.get("/incidents/{incident_id}")
def get_regional_incident_detail(
    incident_id: int,
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db)],
):
    """Fetch a single incident detail, scoped to user's region."""
    region_id = user["assigned_region_id"]

    row = db.execute(
        text("""
            SELECT fi.incident_id, fi.verification_status, fi.created_at,
                   fi.region_id, fi.encoder_id
            FROM wims.fire_incidents fi
            WHERE fi.incident_id = :iid AND fi.region_id = :rid AND fi.is_archived = FALSE
        """),
        {"iid": incident_id, "rid": region_id},
    ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Incident not found in your region")

    # Fetch nonsensitive
    ns = db.execute(
        text("SELECT * FROM wims.incident_nonsensitive_details WHERE incident_id = :iid"),
        {"iid": incident_id},
    ).fetchone()

    # Fetch sensitive
    sd = db.execute(
        text("SELECT * FROM wims.incident_sensitive_details WHERE incident_id = :iid"),
        {"iid": incident_id},
    ).fetchone()

    def row_to_dict(r, keys=None):
        if r is None:
            return {}
        if keys:
            return {k: r[i] for i, k in enumerate(keys)}
        return dict(r._mapping) if hasattr(r, '_mapping') else {}

    return {
        "incident_id": row[0],
        "verification_status": row[1],
        "created_at": row[2].isoformat() if row[2] else None,
        "region_id": row[3],
        "nonsensitive": row_to_dict(ns),
        "sensitive": row_to_dict(sd),
    }


@router.get("/stats", response_model=RegionalStatsResponse)
def get_regional_stats(
    user: Annotated[dict, Depends(get_regional_encoder)],
    db: Annotated[Session, Depends(get_db)],
):
    """Quick summary stats scoped to the user's region."""
    region_id = user["assigned_region_id"]

    total = db.execute(
        text("SELECT COUNT(*) FROM wims.fire_incidents WHERE region_id = :rid AND is_archived = FALSE"),
        {"rid": region_id},
    ).scalar() or 0

    by_cat_rows = db.execute(
        text("""
            SELECT nd.general_category, COUNT(*) as cnt
            FROM wims.fire_incidents fi
            JOIN wims.incident_nonsensitive_details nd ON nd.incident_id = fi.incident_id
            WHERE fi.region_id = :rid AND fi.is_archived = FALSE
            GROUP BY nd.general_category
            ORDER BY cnt DESC
        """),
        {"rid": region_id},
    ).fetchall()

    by_alarm_rows = db.execute(
        text("""
            SELECT nd.alarm_level, COUNT(*) as cnt
            FROM wims.fire_incidents fi
            JOIN wims.incident_nonsensitive_details nd ON nd.incident_id = fi.incident_id
            WHERE fi.region_id = :rid AND fi.is_archived = FALSE
            GROUP BY nd.alarm_level
            ORDER BY cnt DESC
        """),
        {"rid": region_id},
    ).fetchall()

    by_status_rows = db.execute(
        text("""
            SELECT verification_status, COUNT(*) as cnt
            FROM wims.fire_incidents
            WHERE region_id = :rid AND is_archived = FALSE
            GROUP BY verification_status
            ORDER BY cnt DESC
        """),
        {"rid": region_id},
    ).fetchall()

    return RegionalStatsResponse(
        total_incidents=total,
        by_category=[{"category": r[0], "count": r[1]} for r in by_cat_rows],
        by_alarm_level=[{"alarm_level": r[0], "count": r[1]} for r in by_alarm_rows],
        by_status=[{"status": r[0], "count": r[1]} for r in by_status_rows],
    )
